
import os
import pandas as pd

from werkzeug.utils import secure_filename

from services.planning.planning_header_service import PlanningHeaderService
from services.planning.planning_detail_service import PlanningDetailService
from models.planning_header import PlanningHeader
from models.planning_detail import PlanningDetail
from models.kategori import Kategori
from utils.db import db

class PlanningUploadService:

    UPLOAD_FOLDER = os.path.join(
        os.getcwd(),
        "uploads",
        "planning"
    )
    
    @staticmethod
    def ensure_upload_folder():
        if not os.path.exists(PlanningUploadService.UPLOAD_FOLDER):
            os.makedirs(PlanningUploadService.UPLOAD_FOLDER)
    @staticmethod
    def upload_planning(file,periode,user_id):
        if not file:
            return{
                "success":False,
                "message":"file wajib diisi"
            },400

        filename= secure_filename(file.filename)

        if filename =="":
            return{
                "success":False,
                "message":"nama file wajib diisi"
            },400
        allowed_extension ={
            "xls",
            "xlsx"
        }
        ext = filename.rsplit(".", 1)[1].lower()

        if ext not in allowed_extension:
            return{
                "success":False,
                "message":"extensi file tidak didukung"
            },400

        PlanningUploadService.ensure_upload_folder()

        file_path = os.path.join(
            PlanningUploadService.UPLOAD_FOLDER,
            filename
        )
        file.save(file_path)

    @staticmethod
    def _read_planning_df(file_path):
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names

        # Deteksi apakah file ini merupakan Multi-Sheet Workbook resmi Komite Review Budget PT SAI
        committee_forms = [s for s in sheet_names if any(k in s for k in ("Form E-1", "Form E-9", "Form I-1", "Form E-6", "Form E-7"))]

        if committee_forms:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            records = []
            month_map = {
                'jan': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr', 'may': 'May', 'jun': 'Jun',
                'jul': 'Jul', 'aug': 'Aug', 'sep': 'Sep', 'okt': 'Oct', 'oct': 'Oct', 'nov': 'Nov', 'des': 'Dec', 'dec': 'Dec'
            }

            # 1. Parse Form E-1 (Consumables QC/Pabrik)
            if 'Form E-1' in wb.sheetnames:
                ws = wb['Form E-1']
                month_cols = {}
                for c in range(8, ws.max_column + 1):
                    v7 = str(ws.cell(7, c).value or '').strip().lower()
                    v8 = str(ws.cell(8, c).value or '').strip().lower()
                    for m_key, m_val in month_map.items():
                        if m_key in v7:
                            if 'amount' in v8 or 'amt' in v8 or not v8:
                                month_cols[c] = m_val
                            elif 'usage' in v8 and c + 1 <= ws.max_column:
                                month_cols[c + 1] = m_val
                for r in range(9, ws.max_row + 1):
                    item = ws.cell(r, 1).value
                    if not item or 'total' in str(item).lower():
                        continue
                    item_str = str(item).strip()
                    for c, m_val in month_cols.items():
                        amt = ws.cell(r, c).value
                        if isinstance(amt, (int, float)) and amt > 0:
                            records.append({'month': m_val, 'form': 'E-1', 'item': item_str, 'planning_amount': float(amt), 'remarks': ''})

            # 2. Parse Form E-9 (Calibration & Mtc CF)
            sheet_e9 = 'Form E-9 (2)' if 'Form E-9 (2)' in wb.sheetnames else ('Form E-9' if 'Form E-9' in wb.sheetnames else None)
            if sheet_e9:
                ws = wb[sheet_e9]
                month_cols = {}
                for c in range(7, ws.max_column + 1):
                    v8 = str(ws.cell(8, c).value or '').strip().lower()
                    for m_key, m_val in month_map.items():
                        if m_key == v8 or v8.startswith(m_key):
                            month_cols[c] = m_val
                for r in range(9, ws.max_row + 1):
                    item = ws.cell(r, 2).value
                    code = ws.cell(r, 3).value
                    if not item or 'total' in str(item).lower():
                        continue
                    item_str = str(item).strip()
                    if code and str(code).strip():
                        item_str = f"{item_str} ({str(code).strip()})"
                    for c, m_val in month_cols.items():
                        amt = ws.cell(r, c).value
                        if isinstance(amt, (int, float)) and amt > 0:
                            records.append({'month': m_val, 'form': 'E-9', 'item': item_str, 'planning_amount': float(amt), 'remarks': ''})

            # 3. Parse Form I-1 (Investment Asset CAPEX)
            if 'Form I-1' in wb.sheetnames:
                ws = wb['Form I-1']
                month_cols = {}
                for c in range(8, ws.max_column + 1):
                    v7 = str(ws.cell(7, c).value or '').strip().lower()
                    v8 = str(ws.cell(8, c).value or '').strip().lower()
                    for m_key, m_val in month_map.items():
                        if m_key == v8 or v8.startswith(m_key) or m_key == v7 or v7.startswith(m_key):
                            month_cols[c] = m_val
                for r in range(9, ws.max_row + 1):
                    item = ws.cell(r, 2).value
                    if not item or 'total' in str(item).lower():
                        continue
                    item_str = str(item).strip()
                    code_val = str(ws.cell(r, 1).value or '').strip()
                    for c, m_val in month_cols.items():
                        amt = ws.cell(r, c).value
                        if isinstance(amt, (int, float)) and amt > 0:
                            records.append({'month': m_val, 'form': 'I-1', 'item': item_str, 'planning_amount': float(amt), 'remarks': code_val})

            if records:
                return pd.DataFrame(records)

        # Fallback: Single-sheet flat planning template
        sheet_name = None
        for name in ["Budget Planning Detail", "Planning", "Sheet1"]:
            if name in sheet_names:
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = sheet_names[0]

        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df.columns = [
            str(col).strip().lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
            for col in df.columns
        ]

        # Column aliases mapping
        col_aliases = {
            "item_description": "item",
            "planning_amount_idr": "planning_amount",
            "remarks_actual_item": "remarks",
            "category": "form"
        }
        df.rename(columns=col_aliases, inplace=True)

        # Filter out total row if present
        if "item" in df.columns:
            df = df[df["item"].notna() & (~df["item"].astype(str).str.upper().str.contains("TOTAL BUDGET"))]
        elif "form" in df.columns:
            df = df[df["form"].notna() & (~df["form"].astype(str).str.upper().str.contains("TOTAL"))]

        return df

    @staticmethod
    def upload_planning(file, periode, user_id):
        if not file:
            return {
                "success": False,
                "message": "file wajib diisi"
            }, 400

        filename = secure_filename(file.filename)
        if filename == "":
            return {
                "success": False,
                "message": "nama file wajib diisi"
            }, 400

        allowed_extension = {"xls", "xlsx"}
        ext = filename.rsplit(".", 1)[-1].lower()
        if ext not in allowed_extension:
            return {
                "success": False,
                "message": "extensi file tidak didukung"
            }, 400

        PlanningUploadService.ensure_upload_folder()
        file_path = os.path.join(PlanningUploadService.UPLOAD_FOLDER, filename)
        file.save(file_path)

        try:
            df = PlanningUploadService._read_planning_df(file_path)
        except Exception as read_err:
            return {
                "success": False,
                "message": f"Gagal membaca file Excel: {str(read_err)}"
            }, 400

        required_columns = ["month", "form", "item", "planning_amount"]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return {
                "success": False,
                "message": "Missing required columns: " + ", ".join(missing_columns)
            }, 400

        # Cek apakah header periode sudah ada (Re-use existing header or create new)
        existing_header = PlanningHeader.query.filter_by(periode=periode).first()
        if existing_header:
            existing_header.filename = filename
            existing_header.user_id = user_id
            existing_header.status = "UPLOADING"
            planning_header_id = existing_header.id
            db.session.commit()
        else:
            header_resp, status_code = PlanningHeaderService.create_planning_header({
                "periode": periode,
                "user_id": user_id,
                "filename": filename
            })
            if not header_resp.get("success"):
                return header_resp, status_code
            planning_header_id = header_resp["data"]["id"]
            db.session.commit()

        from flask import current_app
        import threading
        
        app = current_app._get_current_object()
        
        # Start background thread
        thread = threading.Thread(
            target=PlanningUploadService._process_excel_background,
            args=(app, file_path, planning_header_id)
        )
        thread.start()

        return {
            "success": True,
            "message": "File sedang diproses di background",
            "data": {
                "planning_header_id": planning_header_id,
                "status": "UPLOADING"
            }
        }, 202

    @staticmethod
    def _process_excel_background(app, file_path, planning_header_id):
        with app.app_context():
            try:
                df = PlanningUploadService._read_planning_df(file_path)

                def get_clean_str(row, col):
                    if col not in row:
                        return ""
                    val = row[col]
                    if isinstance(val, pd.Series):
                        non_na = val.dropna()
                        val = non_na.iloc[0] if not non_na.empty else ""
                    if val is None or pd.isna(val):
                        return ""
                    return str(val).strip()

                def get_clean_float(row, col):
                    if col not in row:
                        return 0.0
                    val = row[col]
                    if isinstance(val, pd.Series):
                        non_na = val.dropna()
                        val = non_na.iloc[0] if not non_na.empty else 0.0
                    if val is None or pd.isna(val):
                        return 0.0
                    try:
                        clean_num = str(val).replace(",", "").strip()
                        return float(clean_num)
                    except Exception:
                        return 0.0

                # Pre-fetch existing planning details for UPSERT (anti-double)
                existing_details = PlanningDetail.query.filter_by(planning_header_id=planning_header_id).all()
                existing_map = {
                    (str(d.month or "").strip().lower(), str(d.item or "").strip().lower()): d
                    for d in existing_details
                }

                for _, row in df.iterrows():
                    form_val = get_clean_str(row, "form")
                    kategori = None
                    if form_val:
                        kategori = Kategori.query.filter(
                            (Kategori.kode.ilike(f"%{form_val}%")) | (Kategori.nama.ilike(f"%{form_val}%"))
                        ).first()

                    if kategori:
                        kategori_id = kategori.id
                    else:
                        item_lower = get_clean_str(row, "item").lower()
                        if item_lower.startswith("kalibrasi") or item_lower == "preventive c/f":
                            kategori_id = 2  # E-9 Calibration & Mtc CF
                        elif any(kw in item_lower for kw in ("cutting machine", "laptop", "lemari", "shredder", "torque", "ojiyas", "ring gauge", "feeler gauge", "filler gauge", "pop nut", "komputer", "machine")):
                            kategori_id = 3  # I-1 Investment Asset CAPEX
                        else:
                            kategori_id = 1  # E-1 Consumable OPEX

                    month_val = get_clean_str(row, "month") or None
                    item_val = get_clean_str(row, "item")
                    plan_amount = get_clean_float(row, "planning_amount")
                    remarks_val = get_clean_str(row, "remarks")

                    if not item_val:
                        continue

                    # Check if already exists (UPSERT)
                    key = (str(month_val or "").strip().lower(), item_val.lower())
                    existing_detail = existing_map.get(key)

                    if existing_detail:
                        existing_detail.planning_amount = plan_amount
                        existing_detail.remarks = remarks_val
                        existing_detail.kategori_id = kategori_id
                    else:
                        new_detail = PlanningDetail(
                            planning_header_id=planning_header_id,
                            kategori_id=kategori_id,
                            month=month_val,
                            item=item_val,
                            planning_amount=plan_amount,
                            remarks=remarks_val
                        )
                        db.session.add(new_detail)
                        existing_map[key] = new_detail

                db.session.commit()

                # Auto-sync total anggaran ke tabel Budget per kategori untuk periode ini
                header = db.session.get(PlanningHeader, planning_header_id)
                if header and header.periode:
                    from models.budget import Budget
                    from sqlalchemy import func
                    for kat in Kategori.query.all():
                        tot = db.session.query(func.sum(PlanningDetail.planning_amount)).filter_by(
                            planning_header_id=planning_header_id,
                            kategori_id=kat.id
                        ).scalar() or 0
                        b = Budget.query.filter_by(periode=header.periode, kategori_id=kat.id).first()
                        if b:
                            b.nominal = tot
                        elif tot > 0:
                            b = Budget(periode=header.periode, kategori_id=kat.id, nominal=tot)
                            db.session.add(b)
                    db.session.commit()

                # Ubah status SUCCES dan commit
                PlanningHeaderService.update_status(planning_header_id, "SUCCES", commit=True)

            except Exception as e:
                import traceback
                db.session.rollback()
                print(f"[PlanningUploadService] Error: {e}")
                traceback.print_exc()
                PlanningHeaderService.update_status(planning_header_id, "FAILED", commit=True)
