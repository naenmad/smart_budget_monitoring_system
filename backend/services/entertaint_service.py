import os
import io
import uuid
from datetime import datetime, date
from decimal import Decimal
from PIL import Image, ImageOps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

from models.entertaint_cost import EntertaintCost
from models.entertaint_receipt import EntertaintReceipt
from models.entertaint_cashflow import EntertaintCashflow
from models.entertaint_master import EntertaintMasterItem
from utils.db import db
from sqlalchemy import or_, and_, func, extract


class EntertaintService:
    UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads", "entertaint")
    MAX_IMAGE_DIMENSION = 1920
    WEBP_QUALITY = 80
    MAX_RAW_FILE_SIZE = 15 * 1024 * 1024  # 15 MB raw limit

    @classmethod
    def ensure_upload_folder(cls):
        os.makedirs(cls.UPLOAD_FOLDER, exist_ok=True)

    # -------------------------------------------------------------------------
    # Claim Cost CRUD
    # -------------------------------------------------------------------------
    @classmethod
    def get_all(
        cls,
        page: int = 1,
        per_page: int = 20,
        search: str = "",
        customer: str = "",
        pic: str = "",
        status_pembayaran: str = "",
        status_claim: str = "",
        start_date: str = "",
        end_date: str = "",
        sort_by: str = "tanggal",
        sort_order: str = "desc"
    ):
        query = EntertaintCost.query

        if search:
            s_term = f"%{search.strip()}%"
            query = query.filter(
                or_(
                    EntertaintCost.deskripsi.ilike(s_term),
                    EntertaintCost.customer.ilike(s_term),
                    EntertaintCost.pic_entertaint.ilike(s_term),
                    EntertaintCost.keterangan.ilike(s_term),
                    EntertaintCost.part_no.ilike(s_term),
                    EntertaintCost.part_name.ilike(s_term),
                    EntertaintCost.problem.ilike(s_term),
                    EntertaintCost.place_of_occurrence.ilike(s_term),
                )
            )

        if customer:
            query = query.filter(EntertaintCost.customer.ilike(f"%{customer.strip()}%"))

        if pic:
            query = query.filter(EntertaintCost.pic_entertaint.ilike(f"%{pic.strip()}%"))

        if status_pembayaran:
            query = query.filter(EntertaintCost.status_pembayaran == status_pembayaran)

        if status_claim:
            query = query.filter(EntertaintCost.status_claim == status_claim)

        if start_date:
            try:
                sd = datetime.strptime(start_date, "%Y-%m-%d").date()
                query = query.filter(EntertaintCost.tanggal >= sd)
            except ValueError:
                pass

        if end_date:
            try:
                ed = datetime.strptime(end_date, "%Y-%m-%d").date()
                query = query.filter(EntertaintCost.tanggal <= ed)
            except ValueError:
                pass

        sort_column = getattr(EntertaintCost, sort_by, EntertaintCost.tanggal)
        if sort_order.lower() == "asc":
            query = query.order_by(sort_column.asc(), EntertaintCost.id.asc())
        else:
            query = query.order_by(sort_column.desc(), EntertaintCost.id.desc())

        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        return {
            "success": True,
            "data": [item.to_dict() for item in pagination.items],
            "total": pagination.total,
            "page": pagination.page,
            "per_page": pagination.per_page,
            "pages": pagination.pages,
        }

    @classmethod
    def get_by_id(cls, cost_id: int):
        item = db.session.get(EntertaintCost, cost_id)
        if not item:
            return {"success": False, "message": "Data Entertainment Cost tidak ditemukan"}, 404
        return {"success": True, "data": item.to_dict()}, 200

    @staticmethod
    def normalize_pic_name(raw_text: str) -> str:
        if not raw_text:
            return None
        lines = str(raw_text).replace('\r', '').split('\n')
        first = lines[0].strip()
        for sep in ('/', ' and ', '&', ' Mr. ', ' Mr '):
            if sep in first:
                first = first.split(sep)[0].strip()

        cleaned = first.replace('Mr. ', '').replace('Mr.', '').replace('Mr ', '').strip()
        low = cleaned.lower()

        if 'yayan' in low:
            return 'Yayan Nuryana'
        if 'andi n' in low or 'andi' in low or 'andhi' in low:
            if 'andhi eko' in low:
                return 'Andhi Eko'
            return 'Andi Nurdiyana'
        if 'kirnanto' in low:
            return 'Kirnanto'
        if 'adi p' in low or 'adi priyanto' in low:
            return 'Adi Priyanto'
        if 'misbah' in low or 'ahmad m' in low:
            return 'Ahmad Misbah'
        if 'yosi' in low:
            return 'Yosi Santori'
        if 'randi' in low:
            return 'Randi Nurandi'
        if 'ardi' in low:
            return 'Ardi S.'
        if 'bisri' in low:
            return 'Muhammad Bisri'
        if 'raihan' in low or 'rhaihan' in low:
            return 'Raihan'
        if 'fadhil' in low:
            return 'Fadhilah'
        if 'maskuri' in low:
            return 'Maskuri'
        if 'arif' in low:
            return 'Arif Basuki'
        if 'wisnu' in low:
            return 'Wisnu'
        if 'ridwan' in low:
            return 'Ridwan'
        if 'sandi' in low:
            return 'Sandi'
        if 'joko susilo' in low:
            return 'Joko Susilo'
        if 'whaski' in low:
            return 'Whaski'
        if 'alvito' in low:
            return 'Alvito'
        if 'jaenal' in low:
            return 'Jaenal'
        if 'nita' in low:
            return 'Nita Marketing'
        if 'sukamto' in low:
            return 'Sukamto'
        if 'choliq' in low:
            return 'Choliq Nugroho'
        if 'teerapong' in low:
            return 'Teerapong'
        if 'takahasi' in low:
            return 'Takahashi'

        return cleaned or None

    @classmethod
    def create_cost(cls, data: dict, user_id: int = None):
        if not data.get("tanggal"):
            return {"success": False, "message": "Tanggal entertaint wajib diisi"}, 400

        if not data.get("deskripsi") or not str(data.get("deskripsi")).strip():
            return {"success": False, "message": "Deskripsi entertaint wajib diisi"}, 400

        try:
            if isinstance(data["tanggal"], str):
                tanggal_val = datetime.strptime(data["tanggal"][:10], "%Y-%m-%d").date()
            else:
                tanggal_val = data["tanggal"]
        except Exception as e:
            return {"success": False, "message": f"Format tanggal tidak valid: {str(e)}"}, 400

        def parse_dec(val):
            if val is None or val == "":
                return Decimal("0.00")
            try:
                return Decimal(str(val).replace("Rp", "").replace(",", "").strip())
            except Exception:
                return Decimal("0.00")

        s1 = parse_dec(data.get("struk_1"))
        s2 = parse_dec(data.get("struk_2"))
        s3 = parse_dec(data.get("struk_3"))
        s4 = parse_dec(data.get("struk_4"))
        sum_struk = s1 + s2 + s3 + s4

        total_amount = parse_dec(data.get("total_amount"))
        if total_amount == 0 and sum_struk > 0:
            total_amount = sum_struk

        total_kasbon = parse_dec(data.get("total_kasbon"))

        def parse_opt_date(v):
            if not v:
                return None
            try:
                if isinstance(v, str):
                    return datetime.strptime(v[:10], "%Y-%m-%d").date()
                return v
            except Exception:
                return None

        normalized_pic = cls.normalize_pic_name(data.get("pic_entertaint"))

        cost = EntertaintCost(
            tanggal=tanggal_val,
            deskripsi=str(data.get("deskripsi")).strip(),
            total_amount=total_amount,
            status_pembayaran=data.get("status_pembayaran", "BELUM_DIBAYAR"),
            status_claim=data.get("status_claim", "OPEN"),
            pic_entertaint=normalized_pic,
            customer=str(data.get("customer", "")).strip() or None,
            place_of_occurrence=str(data.get("place_of_occurrence", "")).strip() or None,
            customer_member=str(data.get("customer_member", "")).strip() or None,
            sai_member=str(data.get("sai_member", "")).strip() or None,
            tanggal_kasbon=parse_opt_date(data.get("tanggal_kasbon")),
            total_kasbon=total_kasbon,
            status_kasbon=str(data.get("status_kasbon", "Belum Lunas")).strip(),
            tanggal_closing=parse_opt_date(data.get("tanggal_closing")),
            keterangan=str(data.get("keterangan", "")).strip() or None,
            part_no=str(data.get("part_no", "")).strip() or None,
            part_name=str(data.get("part_name", "")).strip() or None,
            problem=str(data.get("problem", "")).strip() or None,
            problem_maker=str(data.get("problem_maker", "")).strip() or None,
            qty_problem=int(data["qty_problem"]) if data.get("qty_problem") and str(data["qty_problem"]).isdigit() else None,
            struk_1=s1,
            struk_2=s2,
            struk_3=s3,
            struk_4=s4,
            created_by=user_id
        )

        db.session.add(cost)

        # Auto-enroll new Master Data if not already present
        cls.ensure_master_item("CUSTOMER", cost.customer)
        cls.ensure_master_item("PIC", cost.pic_entertaint)
        cls.ensure_master_item("PLACE", cost.place_of_occurrence)

        db.session.commit()

        return {"success": True, "message": "Entertainment Cost berhasil dicatat", "data": cost.to_dict()}, 201

    @classmethod
    def ensure_master_item(cls, category: str, name: str):
        if not name or not str(name).strip():
            return
        cleaned_name = str(name).strip()
        cat = category.upper()
        existing = EntertaintMasterItem.query.filter(
            EntertaintMasterItem.category == cat,
            func.lower(EntertaintMasterItem.name) == cleaned_name.lower()
        ).first()
        if not existing:
            new_item = EntertaintMasterItem(category=cat, name=cleaned_name, is_active=True)
            db.session.add(new_item)
        elif not existing.is_active:
            existing.is_active = True

    @classmethod
    def update_cost(cls, cost_id: int, data: dict):
        cost = db.session.get(EntertaintCost, cost_id)
        if not cost:
            return {"success": False, "message": "Data tidak ditemukan"}, 404

        if "tanggal" in data and data["tanggal"]:
            try:
                if isinstance(data["tanggal"], str):
                    cost.tanggal = datetime.strptime(data["tanggal"][:10], "%Y-%m-%d").date()
                else:
                    cost.tanggal = data["tanggal"]
            except Exception as e:
                return {"success": False, "message": f"Format tanggal tidak valid: {str(e)}"}, 400

        if "deskripsi" in data:
            cost.deskripsi = str(data["deskripsi"]).strip()

        def parse_dec(val):
            if val is None or val == "":
                return Decimal("0.00")
            try:
                return Decimal(str(val).replace("Rp", "").replace(",", "").strip())
            except Exception:
                return Decimal("0.00")

        if "struk_1" in data:
            cost.struk_1 = parse_dec(data["struk_1"])
        if "struk_2" in data:
            cost.struk_2 = parse_dec(data["struk_2"])
        if "struk_3" in data:
            cost.struk_3 = parse_dec(data["struk_3"])
        if "struk_4" in data:
            cost.struk_4 = parse_dec(data["struk_4"])

        if "total_amount" in data:
            cost.total_amount = parse_dec(data["total_amount"])

        if "status_pembayaran" in data:
            cost.status_pembayaran = data["status_pembayaran"]

        if "status_claim" in data:
            cost.status_claim = data["status_claim"]

        if "pic_entertaint" in data:
            cost.pic_entertaint = cls.normalize_pic_name(data["pic_entertaint"])

        if "customer" in data:
            cost.customer = str(data["customer"]).strip() or None

        if "place_of_occurrence" in data:
            cost.place_of_occurrence = str(data["place_of_occurrence"]).strip() or None

        if "customer_member" in data:
            cost.customer_member = str(data["customer_member"]).strip() or None

        if "sai_member" in data:
            cost.sai_member = str(data["sai_member"]).strip() or None

        def parse_opt_date(v):
            if not v:
                return None
            try:
                if isinstance(v, str):
                    return datetime.strptime(v[:10], "%Y-%m-%d").date()
                return v
            except Exception:
                return None

        if "tanggal_kasbon" in data:
            cost.tanggal_kasbon = parse_opt_date(data["tanggal_kasbon"])

        if "total_kasbon" in data:
            cost.total_kasbon = parse_dec(data["total_kasbon"])

        if "status_kasbon" in data:
            cost.status_kasbon = str(data["status_kasbon"]).strip()

        if "tanggal_closing" in data:
            cost.tanggal_closing = parse_opt_date(data["tanggal_closing"])

        if "keterangan" in data:
            cost.keterangan = str(data["keterangan"]).strip() or None

        if "part_no" in data:
            cost.part_no = str(data["part_no"]).strip() or None

        if "part_name" in data:
            cost.part_name = str(data["part_name"]).strip() or None

        if "problem" in data:
            cost.problem = str(data["problem"]).strip() or None

        if "problem_maker" in data:
            cost.problem_maker = str(data["problem_maker"]).strip() or None

        if "qty_problem" in data:
            cost.qty_problem = int(data["qty_problem"]) if data.get("qty_problem") and str(data["qty_problem"]).isdigit() else None

        # Auto-enroll master items on update
        cls.ensure_master_item("CUSTOMER", cost.customer)
        cls.ensure_master_item("PIC", cost.pic_entertaint)
        cls.ensure_master_item("PLACE", cost.place_of_occurrence)

        db.session.commit()
        return {"success": True, "message": "Data berhasil diperbarui", "data": cost.to_dict()}, 200

    @classmethod
    def delete_cost(cls, cost_id: int):
        cost = db.session.get(EntertaintCost, cost_id)
        if not cost:
            return {"success": False, "message": "Data tidak ditemukan"}, 404

        for r in cost.receipts:
            try:
                abs_path = os.path.join(cls.UPLOAD_FOLDER, r.file_name)
                if os.path.exists(abs_path):
                    os.remove(abs_path)
            except Exception as err:
                print(f"[EntertaintService] Warning removing file {r.file_name}: {err}")

        db.session.delete(cost)
        db.session.commit()
        return {"success": True, "message": "Data Entertainment Cost beserta lampiran berhasil dihapus"}, 200

    # -------------------------------------------------------------------------
    # Receipt Handling: Auto-Compression & Format Conversion to .webp
    # -------------------------------------------------------------------------
    @classmethod
    def save_and_compress_receipt(cls, file_storage, entertaint_cost_id: int):
        cls.ensure_upload_folder()

        if not file_storage or not file_storage.filename:
            return None, "File kosong atau tidak valid"

        filename = file_storage.filename
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        allowed_extensions = {"jpg", "jpeg", "png", "webp", "bmp", "tiff", "heic"}

        if ext not in allowed_extensions:
            return None, f"Format file .{ext} tidak didukung. Gunakan format gambar (JPG, PNG, WebP)"

        file_bytes = file_storage.read()
        if len(file_bytes) > cls.MAX_RAW_FILE_SIZE:
            return None, f"Ukuran file {filename} melebihi batas maksimal 15MB"

        try:
            img = Image.open(io.BytesIO(file_bytes))
            img = ImageOps.exif_transpose(img)

            if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
                img = img.convert("RGBA")
            else:
                img = img.convert("RGB")

            orig_width, orig_height = img.size
            if orig_width > cls.MAX_IMAGE_DIMENSION or orig_height > cls.MAX_IMAGE_DIMENSION:
                img.thumbnail((cls.MAX_IMAGE_DIMENSION, cls.MAX_IMAGE_DIMENSION), Image.Resampling.LANCZOS)

            final_width, final_height = img.size

            unique_id = uuid.uuid4().hex[:12]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            webp_filename = f"struk_{entertaint_cost_id}_{timestamp}_{unique_id}.webp"
            save_path = os.path.join(cls.UPLOAD_FOLDER, webp_filename)

            img.save(save_path, "WEBP", quality=cls.WEBP_QUALITY, method=6)
            final_file_size = os.path.getsize(save_path)

            receipt = EntertaintReceipt(
                entertaint_cost_id=entertaint_cost_id,
                file_name=webp_filename,
                original_name=filename,
                file_path=f"uploads/entertaint/{webp_filename}",
                file_size=final_file_size,
                mime_type="image/webp",
                width=final_width,
                height=final_height
            )
            db.session.add(receipt)
            db.session.commit()

            return receipt, None

        except Exception as e:
            return None, f"Gagal memproses gambar {filename}: {str(e)}"

    @classmethod
    def delete_receipt(cls, receipt_id: int):
        receipt = db.session.get(EntertaintReceipt, receipt_id)
        if not receipt:
            return {"success": False, "message": "Lampiran struk tidak ditemukan"}, 404

        try:
            abs_path = os.path.join(cls.UPLOAD_FOLDER, receipt.file_name)
            if os.path.exists(abs_path):
                os.remove(abs_path)
        except Exception as err:
            print(f"[EntertaintService] Warning removing file {receipt.file_name}: {err}")

        db.session.delete(receipt)
        db.session.commit()
        return {"success": True, "message": "Lampiran struk berhasil dihapus"}, 200

    # -------------------------------------------------------------------------
    # Cashflow / Budget Entertaint Handling
    # -------------------------------------------------------------------------
    @classmethod
    def get_cashflows(cls, page: int = 1, per_page: int = 50, flow_type: str = "", search: str = ""):
        query = EntertaintCashflow.query

        if flow_type:
            query = query.filter(EntertaintCashflow.flow_type == flow_type.upper())

        if search:
            s_term = f"%{search.strip()}%"
            query = query.filter(
                or_(
                    EntertaintCashflow.account_deskripsi.ilike(s_term),
                    EntertaintCashflow.doc_no.ilike(s_term),
                    EntertaintCashflow.keterangan.ilike(s_term),
                )
            )

        query = query.order_by(EntertaintCashflow.tanggal.asc(), EntertaintCashflow.id.asc())
        items = query.all()

        total_in = sum(float(x.uang_masuk or 0) for x in items)
        total_out = sum(float(x.uang_keluar or 0) for x in items)
        current_balance = total_in - total_out

        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        return {
            "success": True,
            "data": [x.to_dict() for x in pagination.items],
            "total": pagination.total,
            "page": pagination.page,
            "per_page": pagination.per_page,
            "pages": pagination.pages,
            "summary": {
                "total_uang_masuk": total_in,
                "total_uang_keluar": total_out,
                "current_balance": current_balance,
                "total_transactions": len(items)
            }
        }

    @classmethod
    def create_cashflow(cls, data: dict):
        if not data.get("tanggal"):
            return {"success": False, "message": "Tanggal transaksi kasbon wajib diisi"}, 400

        if not data.get("account_deskripsi") or not str(data.get("account_deskripsi")).strip():
            return {"success": False, "message": "Deskripsi akun / transaksi wajib diisi"}, 400

        try:
            if isinstance(data["tanggal"], str):
                tgl = datetime.strptime(data["tanggal"][:10], "%Y-%m-%d").date()
            else:
                tgl = data["tanggal"]
        except Exception as e:
            return {"success": False, "message": f"Format tanggal tidak valid: {str(e)}"}, 400

        def parse_dec(val):
            if val is None or val == "":
                return Decimal("0.00")
            try:
                return Decimal(str(val).replace("Rp", "").replace(",", "").strip())
            except Exception:
                return Decimal("0.00")

        flow_type = str(data.get("flow_type", "OUT")).upper()
        uang_masuk = parse_dec(data.get("uang_masuk")) if flow_type == "IN" else Decimal("0.00")
        uang_keluar = parse_dec(data.get("uang_keluar")) if flow_type == "OUT" else Decimal("0.00")

        # Hitung running balance terakhir
        last_item = EntertaintCashflow.query.order_by(EntertaintCashflow.id.desc()).first()
        prev_balance = last_item.balance if last_item else Decimal("0.00")
        new_balance = prev_balance + uang_masuk - uang_keluar

        item = EntertaintCashflow(
            doc_no=str(data.get("doc_no", "")).strip() or None,
            tanggal=tgl,
            flow_type=flow_type,
            account_deskripsi=str(data.get("account_deskripsi")).strip(),
            uang_masuk=uang_masuk,
            uang_keluar=uang_keluar,
            balance=new_balance,
            status_entertaint=str(data.get("status_entertaint", "Open")).strip(),
            keterangan=str(data.get("keterangan", "")).strip() or None
        )
        db.session.add(item)
        db.session.commit()

        return {"success": True, "message": "Transaksi kasbon berhasil dicatat", "data": item.to_dict()}, 201

    @classmethod
    def delete_cashflow(cls, cashflow_id: int):
        item = db.session.get(EntertaintCashflow, cashflow_id)
        if not item:
            return {"success": False, "message": "Data kasbon tidak ditemukan"}, 404

        db.session.delete(item)
        db.session.commit()
        return {"success": True, "message": "Transaksi kasbon berhasil dihapus"}, 200

    # -------------------------------------------------------------------------
    # Master Items (Customer, PIC, Place)
    # -------------------------------------------------------------------------
    @classmethod
    def get_master_items(cls):
        items = EntertaintMasterItem.query.filter_by(is_active=True).order_by(EntertaintMasterItem.name.asc()).all()
        customers = [x.to_dict() for x in items if x.category == "CUSTOMER"]
        pics = [x.to_dict() for x in items if x.category == "PIC"]
        places = [x.to_dict() for x in items if x.category == "PLACE"]
        customer_members = [x.to_dict() for x in items if x.category == "CUSTOMER_MEMBER"]

        return {
            "success": True,
            "data": {
                "customers": customers,
                "pics": pics,
                "places": places,
                "customer_members": customer_members,
                "total": len(items)
            }
        }

    @classmethod
    def create_master_item(cls, data: dict):
        category = str(data.get("category", "")).upper()
        name = str(data.get("name", "")).strip()

        if category not in ("CUSTOMER", "PIC", "PLACE", "CUSTOMER_MEMBER"):
            return {"success": False, "message": "Kategori harus CUSTOMER, PIC, PLACE, atau CUSTOMER_MEMBER"}, 400
        if not name:
            return {"success": False, "message": "Nama wajib diisi"}, 400

        existing = EntertaintMasterItem.query.filter_by(category=category, name=name).first()
        if existing:
            if not existing.is_active:
                existing.is_active = True
                db.session.commit()
                return {"success": True, "message": "Master item berhasil diaktifkan kembali", "data": existing.to_dict()}, 200
            return {"success": False, "message": "Data sudah terdaftar"}, 400

        item = EntertaintMasterItem(category=category, name=name, is_active=True)
        db.session.add(item)
        db.session.commit()

        return {"success": True, "message": "Master item berhasil ditambahkan", "data": item.to_dict()}, 201

    @classmethod
    def delete_master_item(cls, item_id: int):
        item = db.session.get(EntertaintMasterItem, item_id)
        if not item:
            return {"success": False, "message": "Data master tidak ditemukan"}, 404

        db.session.delete(item)
        db.session.commit()
        return {"success": True, "message": "Master item berhasil dihapus"}, 200

    # -------------------------------------------------------------------------
    # Dashboard & KPI Statistics
    # -------------------------------------------------------------------------
    @classmethod
    def get_summary_stats(cls, periode: str = None):
        query = EntertaintCost.query
        if periode:
            try:
                year_int = int(periode)
                query = query.filter(extract("year", EntertaintCost.tanggal) == year_int)
            except ValueError:
                pass

        all_items = query.all()

        total_amount = sum(float(c.total_amount or 0) for c in all_items)
        count_total = len(all_items)

        total_lunas = sum(float(c.total_amount or 0) for c in all_items if c.status_pembayaran == "SUDAH_DIBAYAR")
        count_lunas = sum(1 for c in all_items if c.status_pembayaran == "SUDAH_DIBAYAR")

        total_belum_lunas = sum(float(c.total_amount or 0) for c in all_items if c.status_pembayaran == "BELUM_DIBAYAR")
        count_belum_lunas = sum(1 for c in all_items if c.status_pembayaran == "BELUM_DIBAYAR")

        count_open_claim = sum(1 for c in all_items if c.status_claim == "OPEN")
        count_close_claim = sum(1 for c in all_items if c.status_claim == "CLOSE")

        total_kasbon = sum(float(c.total_kasbon or 0) for c in all_items)

        # Cashflow metrics
        cashflow_items = EntertaintCashflow.query.all()
        total_cf_in = sum(float(x.uang_masuk or 0) for x in cashflow_items)
        total_cf_out = sum(float(x.uang_keluar or 0) for x in cashflow_items)
        current_cf_balance = total_cf_in - total_cf_out

        # Available years for filter dropdown
        distinct_years_claims = db.session.query(extract("year", EntertaintCost.tanggal)).distinct().all()
        distinct_years_cf = db.session.query(extract("year", EntertaintCashflow.tanggal)).distinct().all()
        year_set = {int(y[0]) for y in distinct_years_claims if y[0]} | {int(y[0]) for y in distinct_years_cf if y[0]}
        available_years = sorted(list(year_set), reverse=True) if year_set else [datetime.now().year]

        # Customer Breakdown & Percentages
        customer_map = {}
        for c in all_items:
            cust = c.customer or "Lain-lain / Internal"
            if cust not in customer_map:
                customer_map[cust] = {"total": 0.0, "count": 0}
            customer_map[cust]["total"] += float(c.total_amount or 0)
            customer_map[cust]["count"] += 1

        customer_distribution = sorted(
            [
                {
                    "customer": k,
                    "total": v["total"],
                    "count": v["count"],
                    "percentage": round((v["total"] / total_amount * 100) if total_amount > 0 else 0, 1),
                    "avg_per_event": round(v["total"] / v["count"] if v["count"] > 0 else 0, 2)
                }
                for k, v in customer_map.items()
            ],
            key=lambda x: x["total"],
            reverse=True
        )

        top_customers = customer_distribution[:5]

        # PIC Ranking
        pic_map = {}
        for c in all_items:
            pic_name = c.pic_entertaint or "Tanpa PIC"
            if pic_name not in pic_map:
                pic_map[pic_name] = {"total": 0.0, "count": 0}
            pic_map[pic_name]["total"] += float(c.total_amount or 0)
            pic_map[pic_name]["count"] += 1

        pic_ranking = sorted(
            [
                {
                    "pic": k,
                    "total": v["total"],
                    "count": v["count"],
                    "avg_per_event": round(v["total"] / v["count"] if v["count"] > 0 else 0, 2)
                }
                for k, v in pic_map.items()
            ],
            key=lambda x: x["total"],
            reverse=True
        )

        # Monthly Trends (Claims)
        monthly_map = {m: {"month": m, "total": 0.0, "count": 0} for m in range(1, 13)}
        month_names = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Ags", "Sep", "Okt", "Nov", "Des"]

        for c in all_items:
            if c.tanggal:
                m_idx = c.tanggal.month
                monthly_map[m_idx]["total"] += float(c.total_amount or 0)
                monthly_map[m_idx]["count"] += 1

        monthly_trend = [
            {
                "month_num": m,
                "month_name": month_names[m - 1],
                "total": monthly_map[m]["total"],
                "count": monthly_map[m]["count"],
                "avg_per_event": round(monthly_map[m]["total"] / monthly_map[m]["count"] if monthly_map[m]["count"] > 0 else 0, 2)
            }
            for m in range(1, 13)
        ]

        # Monthly Cashflow Comparison
        cf_monthly_map = {m: {"month": m, "uang_masuk": 0.0, "uang_keluar": 0.0} for m in range(1, 13)}
        for cf in cashflow_items:
            if cf.tanggal:
                m_idx = cf.tanggal.month
                cf_monthly_map[m_idx]["uang_masuk"] += float(cf.uang_masuk or 0)
                cf_monthly_map[m_idx]["uang_keluar"] += float(cf.uang_keluar or 0)

        cashflow_monthly = [
            {
                "month_num": m,
                "month_name": month_names[m - 1],
                "uang_masuk": cf_monthly_map[m]["uang_masuk"],
                "uang_keluar": cf_monthly_map[m]["uang_keluar"],
                "net": cf_monthly_map[m]["uang_masuk"] - cf_monthly_map[m]["uang_keluar"]
            }
            for m in range(1, 13)
        ]

        # Payment and Claim Status Percentages
        lunas_pct = round((total_lunas / total_amount * 100) if total_amount > 0 else 0, 1)
        belum_lunas_pct = round((total_belum_lunas / total_amount * 100) if total_amount > 0 else 0, 1)
        close_pct = round((count_close_claim / count_total * 100) if count_total > 0 else 0, 1)
        open_pct = round((count_open_claim / count_total * 100) if count_total > 0 else 0, 1)

        return {
            "success": True,
            "data": {
                "total_amount": total_amount,
                "count_total": count_total,
                "total_lunas": total_lunas,
                "total_sudah_dibayar": total_lunas,
                "count_lunas": count_lunas,
                "lunas_percentage": lunas_pct,
                "total_belum_lunas": total_belum_lunas,
                "total_belum_dibayar": total_belum_lunas,
                "count_belum_dibayar": count_belum_lunas,
                "belum_lunas_percentage": belum_lunas_pct,
                "count_open_claim": count_open_claim,
                "open_claim_percentage": open_pct,
                "count_close_claim": count_close_claim,
                "close_claim_percentage": close_pct,
                "total_kasbon": total_kasbon,
                "cashflow_in": total_cf_in,
                "cashflow_out": total_cf_out,
                "cashflow_balance": current_cf_balance,
                "top_customers": top_customers,
                "customer_distribution": customer_distribution,
                "pic_ranking": pic_ranking,
                "monthly_trend": monthly_trend,
                "cashflow_monthly": cashflow_monthly,
                "available_years": available_years
            }
        }

    # -------------------------------------------------------------------------
    # Multi-Sheet Excel Export Matching Original Monitoring Workbook
    # -------------------------------------------------------------------------
    @classmethod
    def export_excel(cls, search: str = "", customer: str = "", pic: str = "", status_pembayaran: str = "", status_claim: str = "", start_date: str = "", end_date: str = ""):
        query = EntertaintCost.query

        if search:
            s_term = f"%{search.strip()}%"
            query = query.filter(
                or_(
                    EntertaintCost.deskripsi.ilike(s_term),
                    EntertaintCost.customer.ilike(s_term),
                    EntertaintCost.pic_entertaint.ilike(s_term),
                    EntertaintCost.keterangan.ilike(s_term),
                    EntertaintCost.part_no.ilike(s_term),
                    EntertaintCost.part_name.ilike(s_term),
                )
            )

        if customer:
            query = query.filter(EntertaintCost.customer.ilike(f"%{customer.strip()}%"))
        if pic:
            query = query.filter(EntertaintCost.pic_entertaint.ilike(f"%{pic.strip()}%"))
        if status_pembayaran:
            query = query.filter(EntertaintCost.status_pembayaran == status_pembayaran)
        if status_claim:
            query = query.filter(EntertaintCost.status_claim == status_claim)

        if start_date:
            try:
                sd = datetime.strptime(start_date, "%Y-%m-%d").date()
                query = query.filter(EntertaintCost.tanggal >= sd)
            except Exception:
                pass

        if end_date:
            try:
                ed = datetime.strptime(end_date, "%Y-%m-%d").date()
                query = query.filter(EntertaintCost.tanggal <= ed)
            except Exception:
                pass

        items = query.order_by(EntertaintCost.tanggal.asc(), EntertaintCost.id.asc()).all()

        wb = openpyxl.Workbook()

        FONT_FAMILY = "Segoe UI"
        THIN_BORDER = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )
        HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        SUBHEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        TOTAL_FONT = Font(name=FONT_FAMILY, size=10.5, bold=True, color="1E3A8A")

        ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
        ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

        # =====================================================================
        # SHEET 1: Claim Cost Entertaint (Identical to reference structure)
        # =====================================================================
        ws1 = wb.active
        ws1.title = "Claim Cost Entertaint"
        ws1.views.sheetView[0].showGridLines = True

        # Header Titles
        ws1.merge_cells("A1:W1")
        ws1["A1"] = "[MONITORING CLAIM COST ENTERTAINT]"
        ws1["A1"].font = Font(name=FONT_FAMILY, size=14, bold=True, color="1E3A8A")

        total_kasbon_sum = sum(float(x.total_kasbon or 0) for x in items)
        ws1.merge_cells("A2:W2")
        ws1["A2"] = f"Total Kasbon Terdaftar = Rp {total_kasbon_sum:,.2f} | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws1["A2"].font = Font(name=FONT_FAMILY, size=10, italic=True, color="475569")

        # Multi-level Headers: Row 3 & Row 4
        # A: No.
        ws1.merge_cells("A3:A4"); ws1["A3"] = "No."
        # B: Tanggal Kasbon
        ws1.merge_cells("B3:B4"); ws1["B3"] = "Tanggal Kasbon"
        # C: Total Kasbon
        ws1.merge_cells("C3:C4"); ws1["C3"] = "Total Kasbon"
        # D: PIC Entertaint
        ws1.merge_cells("D3:D4"); ws1["D3"] = "PIC Entertaint"
        # E: Tanggal Entertaint
        ws1.merge_cells("E3:E4"); ws1["E3"] = "Tanggal Entertaint"
        # F: Customer
        ws1.merge_cells("F3:F4"); ws1["F3"] = "Customer"
        # G: Deskripsi Entertaint
        ws1.merge_cells("G3:G4"); ws1["G3"] = "Deskripsi Entertaint"
        # H: Customer Member
        ws1.merge_cells("H3:H4"); ws1["H3"] = "Customer Member"
        # I: SAI Member
        ws1.merge_cells("I3:I4"); ws1["I3"] = "SAI Member"

        # J-O: Problem Information (Row 3 merged, Row 4 sub-headers)
        ws1.merge_cells("J3:O3"); ws1["J3"] = "Problem Information"
        ws1["J4"] = "Part No."
        ws1["K4"] = "Part Name"
        ws1["L4"] = "Problem"
        ws1["M4"] = "Place / Lokasi"
        ws1["N4"] = "Qty."
        ws1["O4"] = "Problem Maker"

        # P-S: Struk Entertaint (Row 3 merged, Row 4 sub-headers #1 - #4)
        ws1.merge_cells("P3:S3"); ws1["P3"] = "Struk Entertaint"
        ws1["P4"] = "#1"
        ws1["Q4"] = "#2"
        ws1["R4"] = "#3"
        ws1["S4"] = "#4"

        # T: Total Struk
        ws1.merge_cells("T3:T4"); ws1["T3"] = "Total Struk"
        # U: Status Kasbon (Lunas / Belum Lunas)
        ws1.merge_cells("U3:U4"); ws1["U3"] = "Status Kasbon\n(Lunas / Belum Lunas)"
        # V: Status Kasbon ke Marketing (Tanggal Closing)
        ws1["V3"] = "Status Kasbon ke Marketing"
        ws1["V4"] = "Tanggal Closing"
        # W: Status Claim (Close / Open)
        ws1["W3"] = "Status Claim"
        ws1["W4"] = "(Close / Open)"
        # X: Keterangan
        ws1.merge_cells("X3:X4"); ws1["X3"] = "Keterangan"

        # Apply header styling
        for r_idx in (3, 4):
            ws1.row_dimensions[r_idx].height = 24
            for c_idx in range(1, 25):
                cell = ws1.cell(row=r_idx, column=c_idx)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL if r_idx == 3 else SUBHEADER_FILL
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER

        # Data Rows
        curr_row = 5
        tot_struk_sum = 0
        tot_s1_sum = 0
        tot_s2_sum = 0
        tot_s3_sum = 0
        tot_s4_sum = 0

        for idx, item in enumerate(items, 1):
            ws1.row_dimensions[curr_row].height = 22
            is_zebra = (idx % 2 == 0)

            s1 = float(item.struk_1 or 0)
            s2 = float(item.struk_2 or 0)
            s3 = float(item.struk_3 or 0)
            s4 = float(item.struk_4 or 0)
            tot_amt = float(item.total_amount or 0)
            kasbon_amt = float(item.total_kasbon or 0)

            tot_s1_sum += s1
            tot_s2_sum += s2
            tot_s3_sum += s3
            tot_s4_sum += s4
            tot_struk_sum += tot_amt

            row_data = [
                idx,
                item.tanggal_kasbon.strftime("%d-%m-%Y") if item.tanggal_kasbon else "-",
                kasbon_amt,
                item.pic_entertaint or "-",
                item.tanggal.strftime("%d-%m-%Y") if item.tanggal else "-",
                item.customer or "-",
                item.deskripsi,
                item.customer_member or "-",
                item.sai_member or "-",
                item.part_no or "-",
                item.part_name or "-",
                item.problem or "-",
                item.place_of_occurrence or "-",
                item.qty_problem if item.qty_problem is not None else "-",
                item.problem_maker or "-",
                s1 if s1 > 0 else None,
                s2 if s2 > 0 else None,
                s3 if s3 > 0 else None,
                s4 if s4 > 0 else None,
                tot_amt,
                item.status_kasbon or ("Lunas" if item.status_pembayaran == "SUDAH_DIBAYAR" else "Belum Lunas"),
                item.tanggal_closing.strftime("%d-%m-%Y") if item.tanggal_closing else "-",
                item.status_claim or "OPEN",
                item.keterangan or "-"
            ]

            for c_idx, val in enumerate(row_data, 1):
                cell = ws1.cell(row=curr_row, column=c_idx, value=val)
                cell.font = Font(name=FONT_FAMILY, size=9.5)
                cell.border = THIN_BORDER
                if is_zebra:
                    cell.fill = ZEBRA_FILL

                # Formats
                if c_idx in [1, 2, 5, 14, 21, 22, 23]:
                    cell.alignment = ALIGN_CENTER
                elif c_idx in [3, 16, 17, 18, 19, 20]:
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = ALIGN_LEFT

            curr_row += 1

        # Summary Row Sheet 1
        ws1.row_dimensions[curr_row].height = 26
        ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
        tot_label = ws1.cell(row=curr_row, column=1, value="TOTAL")
        tot_label.font = TOTAL_FONT; tot_label.alignment = ALIGN_CENTER; tot_label.fill = TOTAL_FILL

        for c_idx in range(1, 25):
            c = ws1.cell(row=curr_row, column=c_idx)
            c.border = THIN_BORDER; c.fill = TOTAL_FILL

        c_kb = ws1.cell(row=curr_row, column=3, value=total_kasbon_sum)
        c_kb.font = TOTAL_FONT; c_kb.alignment = ALIGN_RIGHT; c_kb.number_format = "#,##0"

        c_s1 = ws1.cell(row=curr_row, column=16, value=tot_s1_sum)
        c_s1.font = TOTAL_FONT; c_s1.alignment = ALIGN_RIGHT; c_s1.number_format = "#,##0"
        c_s2 = ws1.cell(row=curr_row, column=17, value=tot_s2_sum)
        c_s2.font = TOTAL_FONT; c_s2.alignment = ALIGN_RIGHT; c_s2.number_format = "#,##0"
        c_s3 = ws1.cell(row=curr_row, column=18, value=tot_s3_sum)
        c_s3.font = TOTAL_FONT; c_s3.alignment = ALIGN_RIGHT; c_s3.number_format = "#,##0"
        c_s4 = ws1.cell(row=curr_row, column=19, value=tot_s4_sum)
        c_s4.font = TOTAL_FONT; c_s4.alignment = ALIGN_RIGHT; c_s4.number_format = "#,##0"

        c_tot = ws1.cell(row=curr_row, column=20, value=tot_struk_sum)
        c_tot.font = TOTAL_FONT; c_tot.alignment = ALIGN_RIGHT; c_tot.number_format = "#,##0"

        for col in ws1.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(len(str(cell.value or "")) for cell in col if cell.row > 2)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 11)

        # =====================================================================
        # SHEET 2: Budget Entertaint (Cash In, Cash Out, Running Balance)
        # =====================================================================
        ws2 = wb.create_sheet(title="Budget Entertaint")
        ws2.views.sheetView[0].showGridLines = True

        ws2.merge_cells("A1:H1")
        ws2["A1"] = "[STATUS BUDGET ENTERTAINT / KAS KASBON]"
        ws2["A1"].font = Font(name=FONT_FAMILY, size=13, bold=True, color="1E3A8A")

        headers_b = ["No.", "Doc No.", "Date", "Account / Deskripsi", "Uang Masuk (ke QC)", "Uang Keluar (ke PIC)", "Balance (Saldo)", "Status Entertaint", "Keterangan"]
        ws2.row_dimensions[3].height = 26
        for c_idx, h in enumerate(headers_b, 1):
            c = ws2.cell(row=3, column=c_idx, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = ALIGN_CENTER; c.border = THIN_BORDER

        cashflow_rows = EntertaintCashflow.query.order_by(EntertaintCashflow.tanggal.asc(), EntertaintCashflow.id.asc()).all()
        r2_curr = 4
        tot_b_in = 0
        tot_b_out = 0

        for b_idx, cf in enumerate(cashflow_rows, 1):
            ws2.row_dimensions[r2_curr].height = 21
            is_zebra = (b_idx % 2 == 0)
            u_in = float(cf.uang_masuk or 0)
            u_out = float(cf.uang_keluar or 0)
            tot_b_in += u_in
            tot_b_out += u_out

            b_row = [
                b_idx,
                cf.doc_no or "-",
                cf.tanggal.strftime("%d-%m-%Y") if cf.tanggal else "-",
                cf.account_deskripsi,
                u_in if u_in > 0 else None,
                u_out if u_out > 0 else None,
                float(cf.balance or 0),
                cf.status_entertaint or "Open",
                cf.keterangan or "-"
            ]

            for c_idx, val in enumerate(b_row, 1):
                cell = ws2.cell(row=r2_curr, column=c_idx, value=val)
                cell.font = Font(name=FONT_FAMILY, size=9.5)
                cell.border = THIN_BORDER
                if is_zebra:
                    cell.fill = ZEBRA_FILL

                if c_idx in [1, 2, 3, 8]:
                    cell.alignment = ALIGN_CENTER
                elif c_idx in [5, 6, 7]:
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = ALIGN_LEFT

            r2_curr += 1

        # Summary Row Sheet 2
        ws2.row_dimensions[r2_curr].height = 26
        ws2.merge_cells(start_row=r2_curr, start_column=1, end_row=r2_curr, end_column=4)
        lbl_b = ws2.cell(row=r2_curr, column=1, value="TOTAL / SISA BALANCE")
        lbl_b.font = TOTAL_FONT; lbl_b.alignment = ALIGN_CENTER; lbl_b.fill = TOTAL_FILL

        for c_idx in range(1, 10):
            ws2.cell(row=r2_curr, column=c_idx).border = THIN_BORDER
            ws2.cell(row=r2_curr, column=c_idx).fill = TOTAL_FILL

        c_bin = ws2.cell(row=r2_curr, column=5, value=tot_b_in)
        c_bin.font = TOTAL_FONT; c_bin.alignment = ALIGN_RIGHT; c_bin.number_format = "#,##0"
        c_bout = ws2.cell(row=r2_curr, column=6, value=tot_b_out)
        c_bout.font = TOTAL_FONT; c_bout.alignment = ALIGN_RIGHT; c_bout.number_format = "#,##0"
        c_bbal = ws2.cell(row=r2_curr, column=7, value=(tot_b_in - tot_b_out))
        c_bbal.font = TOTAL_FONT; c_bbal.alignment = ALIGN_RIGHT; c_bbal.number_format = "#,##0"

        for col in ws2.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(len(str(cell.value or "")) for cell in col if cell.row > 2)
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # =====================================================================
        # SHEET 3: Master List (Customer List, PIC Tugas Luar, Place, Contacts)
        # =====================================================================
        ws3 = wb.create_sheet(title="Master List")
        ws3.views.sheetView[0].showGridLines = True

        ws3.merge_cells("A1:D1")
        ws3["A1"] = "[MASTER LIST REFERENSI]"
        ws3["A1"].font = Font(name=FONT_FAMILY, size=13, bold=True, color="1E3A8A")

        headers_m = ["[Customer List]", "[PIC Tugas Luar]", "[Place of Occurrence]", "[Kontak Personil Customer]"]
        ws3.row_dimensions[3].height = 24
        for c_idx, h in enumerate(headers_m, 1):
            c = ws3.cell(row=3, column=c_idx, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = ALIGN_CENTER; c.border = THIN_BORDER

        masters = EntertaintMasterItem.query.filter_by(is_active=True).order_by(EntertaintMasterItem.name.asc()).all()
        cust_list = [x.name for x in masters if x.category == "CUSTOMER"]
        pic_list = [x.name for x in masters if x.category == "PIC"]
        place_list = [x.name for x in masters if x.category == "PLACE"]
        contact_list = [x.name for x in masters if x.category == "CUSTOMER_MEMBER"]

        max_len_m = max(len(cust_list), len(pic_list), len(place_list), len(contact_list), 1)
        r3_curr = 4
        for m_idx in range(max_len_m):
            ws3.row_dimensions[r3_curr].height = 20
            is_zebra = (m_idx % 2 == 0)

            c_val = cust_list[m_idx] if m_idx < len(cust_list) else ""
            p_val = pic_list[m_idx] if m_idx < len(pic_list) else ""
            pl_val = place_list[m_idx] if m_idx < len(place_list) else ""
            ct_val = contact_list[m_idx] if m_idx < len(contact_list) else ""

            for col_idx, val in enumerate([c_val, p_val, pl_val, ct_val], 1):
                cell = ws3.cell(row=r3_curr, column=col_idx, value=val)
                cell.font = Font(name=FONT_FAMILY, size=9.5)
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_LEFT
                if is_zebra:
                    cell.fill = ZEBRA_FILL

            r3_curr += 1

        for col in ws3.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(len(str(cell.value or "")) for cell in col if cell.row > 2)
            ws3.column_dimensions[col_letter].width = max(max_len + 5, 20)

        # =====================================================================
        # SHEET 4: Lampiran Struk (Receipt Images Embedded)
        # =====================================================================
        ws4 = wb.create_sheet(title="Lampiran Struk")
        ws4.views.sheetView[0].showGridLines = True

        ws4.merge_cells("A1:G1")
        ws4["A1"] = "[LAMPIRAN FOTO STRUK ENTERTAINT]"
        ws4["A1"].font = Font(name=FONT_FAMILY, size=13, bold=True, color="1E3A8A")

        ws4.merge_cells("A2:G2")
        ws4["A2"] = f"Dokumentasi foto struk fisik klaim entertainment & operasional QA | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws4["A2"].font = Font(name=FONT_FAMILY, size=9.5, italic=True, color="64748B")

        r4_curr = 4
        has_any_image = False

        for item in items:
            receipts = item.receipts if hasattr(item, "receipts") and item.receipts else []
            valid_receipts = []
            for r in receipts:
                abs_p = os.path.join(cls.UPLOAD_FOLDER, r.file_name)
                if os.path.exists(abs_p):
                    valid_receipts.append((r, abs_p))

            if not valid_receipts:
                continue

            has_any_image = True
            ws4.row_dimensions[r4_curr].height = 24
            ws4.merge_cells(start_row=r4_curr, start_column=1, end_row=r4_curr, end_column=6)
            card_title = ws4.cell(
                row=r4_curr,
                column=1,
                value=f"Klaim #{item.id} - {item.tanggal.strftime('%d-%m-%Y') if item.tanggal else '-'} | {item.customer or 'Internal'} (PIC: {item.pic_entertaint or '-'})"
            )
            card_title.font = Font(name=FONT_FAMILY, size=10.5, bold=True, color="FFFFFF")
            card_title.fill = HEADER_FILL
            card_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            r4_curr += 1
            ws4.row_dimensions[r4_curr].height = 20
            ws4.merge_cells(start_row=r4_curr, start_column=1, end_row=r4_curr, end_column=6)
            sub_info = ws4.cell(
                row=r4_curr,
                column=1,
                value=f"Deskripsi: {item.deskripsi} | Total Struk: Rp {float(item.total_amount or 0):,.2f} | Status: {item.status_pembayaran}"
            )
            sub_info.font = Font(name=FONT_FAMILY, size=9.5, color="1E3A8A", bold=True)
            sub_info.fill = TOTAL_FILL
            sub_info.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            r4_curr += 1

            for r_obj, abs_path in valid_receipts:
                try:
                    with Image.open(abs_path) as pil_im:
                        w, h = pil_im.size
                        target_w = 300
                        target_h = int(h * (target_w / w)) if w > 0 else 200

                        png_buf = io.BytesIO()
                        pil_im.convert("RGB").save(png_buf, format="PNG")
                        png_buf.seek(0)

                    ox_img = OpenpyxlImage(png_buf)
                    ox_img.width = target_w
                    ox_img.height = target_h

                    ws4.row_dimensions[r4_curr].height = 18
                    lbl = ws4.cell(row=r4_curr, column=2, value=f"Foto: {r_obj.original_name or r_obj.file_name}")
                    lbl.font = Font(name=FONT_FAMILY, size=9, italic=True, color="475569")
                    r4_curr += 1

                    img_cell_coord = f"B{r4_curr}"
                    ws4.add_image(ox_img, img_cell_coord)

                    needed_rows = max(int(target_h / 24) + 2, 8)
                    for ir in range(needed_rows):
                        ws4.row_dimensions[r4_curr + ir].height = 24
                    r4_curr += needed_rows + 1
                except Exception as img_err:
                    err_cell = ws4.cell(row=r4_curr, column=2, value=f"Gagal memuat foto: {str(img_err)}")
                    err_cell.font = Font(name=FONT_FAMILY, size=9, color="DC2626")
                    r4_curr += 2

            r4_curr += 1

        if not has_any_image:
            ws4.merge_cells("B4:F5")
            c_empty = ws4.cell(
                row=4,
                column=2,
                value="Belum ada foto struk yang diunggah ke sistem web.\nFoto struk yang Anda lampirkan saat input klaim akan otomatis tersemat rapi di lembar ini."
            )
            c_empty.font = Font(name=FONT_FAMILY, size=10, italic=True, color="64748B")
            c_empty.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c_empty.fill = TOTAL_FILL

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # -------------------------------------------------------------------------
    # Bidirectional Excel Import (Lossless & Robust UPSERT)
    # -------------------------------------------------------------------------
    @classmethod
    def import_excel(cls, file_storage, user_id: int = None):
        try:
            wb = openpyxl.load_workbook(file_storage, data_only=True)
        except Exception as e:
            return {"success": False, "message": f"Gagal membaca file Excel: {str(e)}"}, 400

        # Sheet detection: prefer 'Claim Cost Entertaint', or first sheet
        target_sheet_name = None
        for sname in wb.sheetnames:
            if "claim" in sname.lower() or "cost" in sname.lower():
                target_sheet_name = sname
                break
        if not target_sheet_name:
            target_sheet_name = wb.sheetnames[0]

        ws = wb[target_sheet_name]

        def parse_date_safe(val):
            if not val:
                return None
            if isinstance(val, (datetime, date)):
                return val if isinstance(val, date) else val.date()
            if isinstance(val, str):
                val = val.strip()
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
                    try:
                        return datetime.strptime(val[:10], fmt).date()
                    except Exception:
                        pass
            return None

        def parse_dec_safe(val):
            if val is None or val == "":
                return Decimal("0.00")
            try:
                cleaned = str(val).replace("Rp", "").replace(",", "").strip()
                return Decimal(cleaned)
            except Exception:
                return Decimal("0.00")

        def clean_str_safe(val):
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        # Find header row
        header_row = 3
        for r in range(1, min(10, ws.max_row + 1)):
            row_texts = [str(ws.cell(r, c).value or "").lower() for c in range(1, 15)]
            if any("tanggal" in v for v in row_texts) or any("customer" in v for v in row_texts) or any("pic" in v for v in row_texts):
                header_row = r
                break

        created_claims = 0
        updated_claims = 0
        skipped_rows = 0

        # Parse claim rows
        for r in range(header_row + 1, ws.max_row + 1):
            tgl_kb = parse_date_safe(ws.cell(r, 2).value)
            tot_kb = parse_dec_safe(ws.cell(r, 3).value)
            pic_raw = clean_str_safe(ws.cell(r, 4).value)
            tgl_ent = parse_date_safe(ws.cell(r, 5).value)
            cust = clean_str_safe(ws.cell(r, 6).value)
            desc_raw = clean_str_safe(ws.cell(r, 7).value)

            # Skip summary / total / footnote rows
            row_prefix_text = " ".join([str(ws.cell(r, c).value or "") for c in range(1, 8)]).upper()
            if "TOTAL" in row_prefix_text or "SUMMARY" in row_prefix_text or "GRAND" in row_prefix_text:
                continue
            if "LIHAT DOC" in row_prefix_text or "LIHAT DI" in row_prefix_text:
                continue

            cust_mem = clean_str_safe(ws.cell(r, 8).value)
            sai_mem = clean_str_safe(ws.cell(r, 9).value)
            part_no = clean_str_safe(ws.cell(r, 10).value)
            part_name = clean_str_safe(ws.cell(r, 11).value)
            problem = clean_str_safe(ws.cell(r, 12).value)
            place = clean_str_safe(ws.cell(r, 13).value)

            qty_val = None
            try:
                q_raw = ws.cell(r, 14).value
                if q_raw and str(q_raw).isdigit():
                    qty_val = int(q_raw)
            except Exception:
                pass

            problem_maker = clean_str_safe(ws.cell(r, 15).value)
            s1 = parse_dec_safe(ws.cell(r, 16).value)
            s2 = parse_dec_safe(ws.cell(r, 17).value)
            s3 = parse_dec_safe(ws.cell(r, 18).value)
            s4 = parse_dec_safe(ws.cell(r, 19).value)
            tot_s = parse_dec_safe(ws.cell(r, 20).value)

            if tot_s == 0:
                sum_s = s1 + s2 + s3 + s4
                if sum_s > 0:
                    tot_s = sum_s
                elif tot_kb > 0:
                    tot_s = tot_kb

            # Must have a valid date AND either amount > 0 or kasbon > 0 AND (cust or pic)
            if not tgl_kb and not tgl_ent:
                skipped_rows += 1
                continue

            if tot_s == 0 and tot_kb == 0:
                skipped_rows += 1
                continue

            if not cust and not pic_raw:
                skipped_rows += 1
                continue

            primary_date = tgl_kb or tgl_ent or date.today()
            norm_pic = cls.normalize_pic_name(pic_raw)

            # Smart fallback for description
            if desc_raw:
                final_desc = desc_raw
            elif problem:
                final_desc = f"Problem: {problem}"
            elif part_name:
                final_desc = f"Joint check {part_name}"
            elif cust:
                final_desc = f"Customer visit {cust}"
            else:
                final_desc = "Aktivitas Entertaint QA"

            # Status Kasbon & Pembayaran
            status_kb_raw = clean_str_safe(ws.cell(r, 21).value) or ""
            if "lunas" in status_kb_raw.lower() or "paid" in status_kb_raw.lower():
                status_pembayaran = "SUDAH_DIBAYAR"
                status_kasbon = "Lunas"
            else:
                status_pembayaran = "BELUM_DIBAYAR"
                status_kasbon = status_kb_raw or "Belum Lunas"

            tgl_closing = parse_date_safe(ws.cell(r, 22).value)
            status_cl_raw = clean_str_safe(ws.cell(r, 23).value) or ""
            status_claim = "CLOSE" if "close" in status_cl_raw.lower() else "OPEN"
            keterangan = clean_str_safe(ws.cell(r, 24).value)

            # Deduplication / UPSERT check
            existing = None
            if cust and norm_pic and tot_s > 0:
                existing = EntertaintCost.query.filter(
                    EntertaintCost.tanggal == primary_date,
                    EntertaintCost.customer == cust,
                    EntertaintCost.pic_entertaint == norm_pic,
                    EntertaintCost.total_amount == tot_s
                ).first()

            if not existing and cust and tot_s > 0:
                existing = EntertaintCost.query.filter(
                    EntertaintCost.tanggal == primary_date,
                    EntertaintCost.customer == cust,
                    EntertaintCost.total_amount == tot_s
                ).first()

            if existing:
                existing.total_kasbon = tot_kb
                existing.tanggal_kasbon = tgl_kb
                existing.deskripsi = final_desc
                existing.customer_member = cust_mem or existing.customer_member
                existing.sai_member = sai_mem or existing.sai_member
                existing.part_no = part_no or existing.part_no
                existing.part_name = part_name or existing.part_name
                existing.problem = problem or existing.problem
                existing.place_of_occurrence = place or existing.place_of_occurrence
                if qty_val is not None:
                    existing.qty_problem = qty_val
                existing.problem_maker = problem_maker or existing.problem_maker
                existing.struk_1 = s1
                existing.struk_2 = s2
                existing.struk_3 = s3
                existing.struk_4 = s4
                existing.status_pembayaran = status_pembayaran
                existing.status_kasbon = status_kasbon
                existing.tanggal_closing = tgl_closing or existing.tanggal_closing
                existing.status_claim = status_claim
                existing.keterangan = keterangan or existing.keterangan
                updated_claims += 1
            else:
                new_cost = EntertaintCost(
                    tanggal=primary_date,
                    deskripsi=final_desc,
                    total_amount=tot_s,
                    status_pembayaran=status_pembayaran,
                    status_claim=status_claim,
                    pic_entertaint=norm_pic,
                    customer=cust,
                    place_of_occurrence=place,
                    customer_member=cust_mem,
                    sai_member=sai_mem,
                    tanggal_kasbon=tgl_kb,
                    total_kasbon=tot_kb,
                    status_kasbon=status_kasbon,
                    tanggal_closing=tgl_closing,
                    keterangan=keterangan,
                    part_no=part_no,
                    part_name=part_name,
                    problem=problem,
                    problem_maker=problem_maker,
                    qty_problem=qty_val,
                    struk_1=s1,
                    struk_2=s2,
                    struk_3=s3,
                    struk_4=s4,
                    created_by=user_id
                )
                db.session.add(new_cost)
                cls.ensure_master_item("CUSTOMER", cust)
                cls.ensure_master_item("PIC", norm_pic)
                cls.ensure_master_item("PLACE", place)
                created_claims += 1

        # Check if Budget Entertaint sheet exists and import cashflows
        cashflow_synced = 0
        cf_sheet_name = None
        for sname in wb.sheetnames:
            if "budget entertaint" in sname.lower():
                cf_sheet_name = sname
                break

        if cf_sheet_name:
            ws_cf = wb[cf_sheet_name]
            for r in range(3, ws_cf.max_row + 1):
                doc_no = clean_str_safe(ws_cf.cell(r, 6).value)
                cf_date = parse_date_safe(ws_cf.cell(r, 7).value)
                desc = clean_str_safe(ws_cf.cell(r, 8).value)
                masuk = parse_dec_safe(ws_cf.cell(r, 9).value)
                keluar = parse_dec_safe(ws_cf.cell(r, 10).value)
                bal = parse_dec_safe(ws_cf.cell(r, 11).value)
                st = clean_str_safe(ws_cf.cell(r, 12).value) or "Open"
                ket = clean_str_safe(ws_cf.cell(r, 13).value)

                if not desc and masuk == 0 and keluar == 0:
                    continue
                if desc and ("total" in desc.lower() or "saldo" in desc.lower()):
                    continue

                existing_cf = EntertaintCashflow.query.filter(
                    EntertaintCashflow.tanggal == cf_date,
                    EntertaintCashflow.account_deskripsi == desc,
                    EntertaintCashflow.uang_masuk == masuk,
                    EntertaintCashflow.uang_keluar == keluar
                ).first()

                if not existing_cf:
                    new_cf = EntertaintCashflow(
                        doc_no=doc_no,
                        tanggal=cf_date or date.today(),
                        flow_type="IN" if masuk > 0 else "OUT",
                        account_deskripsi=desc or "Transaksi Kasbon",
                        uang_masuk=masuk,
                        uang_keluar=keluar,
                        balance=bal,
                        status_entertaint=st,
                        keterangan=ket
                    )
                    db.session.add(new_cf)
                    cashflow_synced += 1

        db.session.commit()

        return {
            "success": True,
            "message": f"Import Excel berhasil: {created_claims} klaim baru ditambahkan, {updated_claims} diperbarui, {cashflow_synced} arus kas disinkronkan.",
            "data": {
                "created_claims": created_claims,
                "updated_claims": updated_claims,
                "cashflow_synced": cashflow_synced,
                "skipped_rows": skipped_rows,
                "total_rows_parsed": created_claims + updated_claims + skipped_rows
            }
        }, 200
