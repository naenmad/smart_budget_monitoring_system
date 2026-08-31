import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.planning_header import PlanningHeader
from models.planning_detail import PlanningDetail
from models.pr_po_data import PrPoData
from models.kategori import Kategori
from utils.db import db


class ExcelExportService:
    FONT_FAMILY = "Segoe UI"

    # Styling definitions
    HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    SUBTOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    KPI_BG_FILL = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")

    KPI_FONT_TITLE = Font(name=FONT_FAMILY, size=10, bold=True, color="595959")
    KPI_FONT_VAL = Font(name=FONT_FAMILY, size=16, bold=True, color="1F4E78")
    KPI_FONT_SUB = Font(name=FONT_FAMILY, size=9, italic=True, color="7F7F7F")

    TITLE_FONT = Font(name=FONT_FAMILY, size=16, bold=True, color="1F4E78")
    SUBTITLE_FONT = Font(name=FONT_FAMILY, size=10, italic=True, color="595959")
    SECTION_FONT = Font(name=FONT_FAMILY, size=12, bold=True, color="1F4E78")

    REGULAR_FONT = Font(name=FONT_FAMILY, size=10, color="000000")
    BOLD_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="000000")

    THIN_BORDER_GRAY = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    TABLE_HEADER_BORDER = Border(
        left=Side(style="thin", color="1F4E78"),
        right=Side(style="thin", color="1F4E78"),
        top=Side(style="medium", color="1F4E78"),
        bottom=Side(style="medium", color="1F4E78")
    )
    TOTAL_ROW_BORDER = Border(
        top=Side(style="thin", color="1F4E78"),
        bottom=Side(style="double", color="1F4E78")
    )

    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @classmethod
    def generate_planning_excel(cls, periode: str) -> io.BytesIO:
        wb = openpyxl.Workbook()

        # Sheet 1: Executive Summary
        ws_dash = wb.active
        ws_dash.title = "Executive Summary"
        ws_dash.views.sheetView[0].showGridLines = True

        # Sheet 2: Budget Planning Detail
        ws_data = wb.create_sheet(title="Budget Planning Detail")
        ws_data.views.sheetView[0].showGridLines = True

        # Fetch planning data
        header = PlanningHeader.query.filter(
            PlanningHeader.periode == periode,
            PlanningHeader.status.in_(["SUCCESS", "SUCCES"])
        ).first()

        planning_details = []
        if header:
            planning_details = PlanningDetail.query.filter(
                PlanningDetail.planning_header_id == header.id,
                PlanningDetail.status_realisasi != "CANCELLED"
            ).order_by(PlanningDetail.id.asc()).all()

        kategoris = {k.id: k for k in Kategori.query.all()}

        # Headers for Detail Sheet
        detail_headers = [
            "No", "User / Section", "Category", "Form", "Item Description", 
            "Month", "PR No (Req ID)", "Remarks (Actual Item)", "Actual Qty (PR)", 
            "Planning Amount (IDR)", "Status Realisasi", "Matched PO DocNum", 
            "PO Spend Amount (IDR)", "Budget Variance (IDR)"
        ]

        ws_data.row_dimensions[1].height = 28
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.ALIGN_HEADER
            cell.border = cls.TABLE_HEADER_BORDER

        # Fetch matching PR data for each planning item
        for i, pd_item in enumerate(planning_details, 1):
            row_idx = i + 1
            ws_data.row_dimensions[row_idx].height = 20

            kat = kategoris.get(pd_item.kategori_id)
            form_code = kat.kode if kat else "E-1"
            cat_label = "CAPEX" if "CAPEX" in form_code else "OPEX"
            user_sec = "QC ADMIN"

            # Find matching PR
            linked_prs = PrPoData.query.filter(
                PrPoData.planning_detail_id == pd_item.id,
                PrPoData.status_ai != "CANCELLED"
            ).all()

            pr_doc_str = ", ".join([str(p.pr_doc_num) for p in linked_prs if p.pr_doc_num]) or None
            po_doc_str = ", ".join([str(p.po_doc_num) for p in linked_prs if p.po_doc_num]) or ("PR Approved (No PO Yet)" if linked_prs else "")
            po_spend_amt = sum([float(p.total_price or 0.0) for p in linked_prs]) if linked_prs else None
            actual_qty = sum([float(p.qty or 0.0) for p in linked_prs]) if linked_prs else None
            status_real = "Realized (PR Issued)" if linked_prs else "Planned (Pending PR)"

            plan_amt = float(pd_item.planning_amount or 0.0)

            c1 = ws_data.cell(row=row_idx, column=1, value=i)
            c2 = ws_data.cell(row=row_idx, column=2, value=user_sec)
            c3 = ws_data.cell(row=row_idx, column=3, value=cat_label)
            c4 = ws_data.cell(row=row_idx, column=4, value=form_code)
            c5 = ws_data.cell(row=row_idx, column=5, value=pd_item.item)
            c6 = ws_data.cell(row=row_idx, column=6, value=pd_item.month)
            c7 = ws_data.cell(row=row_idx, column=7, value=pr_doc_str)
            c8 = ws_data.cell(row=row_idx, column=8, value=pd_item.remarks or "")
            c9 = ws_data.cell(row=row_idx, column=9, value=actual_qty)
            c10 = ws_data.cell(row=row_idx, column=10, value=plan_amt)
            c11 = ws_data.cell(row=row_idx, column=11, value=status_real)
            c12 = ws_data.cell(row=row_idx, column=12, value=po_doc_str)
            c13 = ws_data.cell(row=row_idx, column=13, value=po_spend_amt)
            c14 = ws_data.cell(row=row_idx, column=14, value=f"=J{row_idx}-M{row_idx}" if po_spend_amt is not None else None)

            c1.alignment = cls.ALIGN_CENTER
            c2.alignment = cls.ALIGN_CENTER
            c3.alignment = cls.ALIGN_CENTER
            c4.alignment = cls.ALIGN_CENTER
            c5.alignment = cls.ALIGN_LEFT
            c6.alignment = cls.ALIGN_CENTER
            c7.alignment = cls.ALIGN_CENTER
            c8.alignment = cls.ALIGN_LEFT
            c9.alignment = cls.ALIGN_RIGHT
            c9.number_format = "#,##0.00" if actual_qty is not None else "General"
            c10.alignment = cls.ALIGN_RIGHT
            c10.number_format = "#,##0"
            c11.alignment = cls.ALIGN_CENTER
            c12.alignment = cls.ALIGN_LEFT
            c13.alignment = cls.ALIGN_RIGHT
            c13.number_format = "#,##0"
            c14.alignment = cls.ALIGN_RIGHT
            c14.number_format = "#,##0"

            fill = cls.ZEBRA_FILL if i % 2 == 0 else None
            for cell in [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14]:
                cell.font = cls.REGULAR_FONT
                cell.border = cls.THIN_BORDER_GRAY
                if fill:
                    cell.fill = fill

        total_row_idx = len(planning_details) + 2
        ws_data.row_dimensions[total_row_idx].height = 24
        c_lbl = ws_data.cell(row=total_row_idx, column=5, value=f"TOTAL BUDGET {periode}")
        c_lbl.font = cls.BOLD_FONT
        c_lbl.alignment = cls.ALIGN_RIGHT

        c_tot_plan = ws_data.cell(row=total_row_idx, column=10, value=f"=SUM(J2:J{max(2, total_row_idx-1)})")
        c_tot_plan.font = cls.BOLD_FONT
        c_tot_plan.alignment = cls.ALIGN_RIGHT
        c_tot_plan.number_format = "#,##0"

        c_tot_po = ws_data.cell(row=total_row_idx, column=13, value=f"=SUM(M2:M{max(2, total_row_idx-1)})")
        c_tot_po.font = cls.BOLD_FONT
        c_tot_po.alignment = cls.ALIGN_RIGHT
        c_tot_po.number_format = "#,##0"

        c_tot_var = ws_data.cell(row=total_row_idx, column=14, value=f"=J{total_row_idx}-M{total_row_idx}")
        c_tot_var.font = cls.BOLD_FONT
        c_tot_var.alignment = cls.ALIGN_RIGHT
        c_tot_var.number_format = "#,##0"

        for col in range(1, 15):
            cell = ws_data.cell(row=total_row_idx, column=col)
            cell.fill = cls.SUBTOTAL_FILL
            cell.border = cls.TOTAL_ROW_BORDER

        ws_data.freeze_panes = "A2"
        ws_data.auto_filter.ref = f"A1:N{max(2, total_row_idx-1)}"

        # Column widths
        for col in ws_data.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                v = str(cell.value or "")
                if cell.number_format and "#,##0" in cell.number_format and isinstance(cell.value, (int, float)):
                    v = f"{cell.value:,.0f}"
                max_len = max(max_len, len(v))
            ws_data.column_dimensions[col_letter].width = max(max_len + 4, 11)
        ws_data.column_dimensions["E"].width = 44
        ws_data.column_dimensions["H"].width = 44
        ws_data.column_dimensions["L"].width = 30

        # Build Executive Summary Sheet
        ws_dash.column_dimensions["A"].width = 3
        ws_dash.column_dimensions["B"].width = 24
        ws_dash.column_dimensions["C"].width = 18
        ws_dash.column_dimensions["D"].width = 18
        ws_dash.column_dimensions["E"].width = 18
        ws_dash.column_dimensions["F"].width = 4
        ws_dash.column_dimensions["G"].width = 24
        ws_dash.column_dimensions["H"].width = 18
        ws_dash.column_dimensions["I"].width = 18

        ws_dash.merge_cells("B2:I2")
        ws_dash["B2"] = f"EXECUTIVE SUMMARY - PLANNING BUDGET {periode}"
        ws_dash["B2"].font = cls.TITLE_FONT
        ws_dash.row_dimensions[2].height = 26

        ws_dash.merge_cells("B3:I3")
        ws_dash["B3"] = f"PT Summit Adyawinsa Indonesia | Year {periode} Smart Budget Monitoring System"
        ws_dash["B3"].font = cls.SUBTITLE_FONT
        ws_dash.row_dimensions[3].height = 18

        kpis = [
            ("B5:C5", "B6:C6", "B7:C7", "TOTAL BUDGET PLANNED", f"=SUM('Budget Planning Detail'!J2:J{max(2, total_row_idx-1)})", f"{len(planning_details)} Planned Line Items"),
            ("D5:E5", "D6:E6", "D7:E7", "BUDGET REALIZED (PR ISSUED)", f"=SUMIFS('Budget Planning Detail'!J2:J{max(2, total_row_idx-1)}, 'Budget Planning Detail'!K2:K{max(2, total_row_idx-1)}, \"Realized (PR Issued)\")", "Items with PR Issued"),
            ("G5:H5", "G6:H6", "G7:H7", "BUDGET REMAINING / UNREALIZED", "=B6-D6", "Items pending PR"),
            ("I5:I5", "I6:I6", "I7:I7", "REALIZATION RATE", "=IF(B6>0, D6/B6, 0)", "% Budget Allocated")
        ]

        for title_range, val_range, sub_range, title, formula, sub in kpis:
            t_start = title_range.split(":")[0]
            v_start = val_range.split(":")[0]
            s_start = sub_range.split(":")[0]

            if ":" in title_range and title_range.split(":")[0] != title_range.split(":")[1]:
                ws_dash.merge_cells(title_range)
                ws_dash.merge_cells(val_range)
                ws_dash.merge_cells(sub_range)

            ws_dash[t_start] = title
            ws_dash[t_start].font = cls.KPI_FONT_TITLE
            ws_dash[t_start].alignment = cls.ALIGN_CENTER
            ws_dash[t_start].fill = cls.KPI_BG_FILL

            ws_dash[v_start] = formula
            ws_dash[v_start].font = cls.KPI_FONT_VAL
            ws_dash[v_start].alignment = cls.ALIGN_CENTER
            ws_dash[v_start].fill = cls.KPI_BG_FILL
            if "%" in sub:
                ws_dash[v_start].number_format = "0.0%"
            else:
                ws_dash[v_start].number_format = "Rp #,##0"

            ws_dash[s_start] = sub
            ws_dash[s_start].font = cls.KPI_FONT_SUB
            ws_dash[s_start].alignment = cls.ALIGN_CENTER
            ws_dash[s_start].fill = cls.KPI_BG_FILL

        # Table 1: Category
        ws_dash["B9"] = "1. REKAPITULASI BERDASARKAN KATEGORI"
        ws_dash["B9"].font = cls.SECTION_FONT

        cat_headers = ["Kategori", "Total Budget (IDR)", "Realized (IDR)", "Remaining (IDR)"]
        for c_i, h in enumerate(cat_headers, 2):
            cell = ws_dash.cell(row=10, column=c_i, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.ALIGN_HEADER
            cell.border = cls.TABLE_HEADER_BORDER

        categories = ["OPEX", "CAPEX"]
        for r_i, cat in enumerate(categories, 11):
            ws_dash.cell(row=r_i, column=2, value=cat).alignment = cls.ALIGN_CENTER
            c_tot = ws_dash.cell(row=r_i, column=3, value=f"=SUMIF('Budget Planning Detail'!$C$2:$C${max(2, total_row_idx-1)}, B{r_i}, 'Budget Planning Detail'!$J$2:$J${max(2, total_row_idx-1)})")
            c_real = ws_dash.cell(row=r_i, column=4, value=f"=SUMIFS('Budget Planning Detail'!$J$2:$J${max(2, total_row_idx-1)}, 'Budget Planning Detail'!$C$2:$C${max(2, total_row_idx-1)}, B{r_i}, 'Budget Planning Detail'!$K$2:$K${max(2, total_row_idx-1)}, \"Realized (PR Issued)\")")
            c_rem = ws_dash.cell(row=r_i, column=5, value=f"=C{r_i}-D{r_i}")
            for cell in [c_tot, c_real, c_rem]:
                cell.number_format = "#,##0"
                cell.alignment = cls.ALIGN_RIGHT
            for c_i in range(2, 6):
                c = ws_dash.cell(row=r_i, column=c_i)
                c.font = cls.REGULAR_FONT
                c.border = cls.THIN_BORDER_GRAY

        ws_dash.cell(row=13, column=2, value="TOTAL").alignment = cls.ALIGN_CENTER
        ws_dash.cell(row=13, column=3, value="=SUM(C11:C12)").number_format = "#,##0"
        ws_dash.cell(row=13, column=4, value="=SUM(D11:D12)").number_format = "#,##0"
        ws_dash.cell(row=13, column=5, value="=SUM(E11:E12)").number_format = "#,##0"
        for c_i in range(2, 6):
            c = ws_dash.cell(row=13, column=c_i)
            c.font = cls.BOLD_FONT
            c.fill = cls.SUBTOTAL_FILL
            c.border = cls.TOTAL_ROW_BORDER

        # Table 2: Form
        ws_dash["G9"] = "2. REKAPITULASI BERDASARKAN FORM"
        ws_dash["G9"].font = cls.SECTION_FONT

        form_headers = ["Form", "Total Budget (IDR)", "Realized (IDR)"]
        for c_i, h in enumerate(form_headers, 7):
            cell = ws_dash.cell(row=10, column=c_i, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.ALIGN_HEADER
            cell.border = cls.TABLE_HEADER_BORDER

        forms = ["E-1 (Operational)", "E-9 (Kalibrasi)", "I-1 (Investment/Capex)"]
        form_codes = ["E-1", "E-9", "I-1"]
        for r_i, (f_lbl, f_code) in enumerate(zip(forms, form_codes), 11):
            ws_dash.cell(row=r_i, column=7, value=f_lbl).alignment = cls.ALIGN_LEFT
            c_tot = ws_dash.cell(row=r_i, column=8, value=f"=SUMIF('Budget Planning Detail'!$D$2:$D${max(2, total_row_idx-1)}, \"{f_code}\", 'Budget Planning Detail'!$J$2:$J${max(2, total_row_idx-1)})")
            c_real = ws_dash.cell(row=r_i, column=9, value=f"=SUMIFS('Budget Planning Detail'!$J$2:$J${max(2, total_row_idx-1)}, 'Budget Planning Detail'!$D$2:$D${max(2, total_row_idx-1)}, \"{f_code}\", 'Budget Planning Detail'!$K$2:$K${max(2, total_row_idx-1)}, \"Realized (PR Issued)\")")
            for cell in [c_tot, c_real]:
                cell.number_format = "#,##0"
                cell.alignment = cls.ALIGN_RIGHT
            for c_i in range(7, 10):
                c = ws_dash.cell(row=r_i, column=c_i)
                c.font = cls.REGULAR_FONT
                c.border = cls.THIN_BORDER_GRAY

        ws_dash.cell(row=14, column=7, value="TOTAL").alignment = cls.ALIGN_LEFT
        ws_dash.cell(row=14, column=8, value="=SUM(H11:H13)").number_format = "#,##0"
        ws_dash.cell(row=14, column=9, value="=SUM(I11:I13)").number_format = "#,##0"
        for c_i in range(7, 10):
            c = ws_dash.cell(row=14, column=c_i)
            c.font = cls.BOLD_FONT
            c.fill = cls.SUBTOTAL_FILL
            c.border = cls.TOTAL_ROW_BORDER

        # Table 3: Month
        ws_dash["B16"] = "3. REKAPITULASI BERDASARKAN BULAN"
        ws_dash["B16"].font = cls.SECTION_FONT

        m_headers = ["Bulan", "Total Budget (IDR)", "Realized (IDR)", "Remaining (IDR)"]
        for c_i, h in enumerate(m_headers, 2):
            cell = ws_dash.cell(row=17, column=c_i, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.ALIGN_HEADER
            cell.border = cls.TABLE_HEADER_BORDER

        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for r_i, m in enumerate(month_names, 18):
            ws_dash.cell(row=r_i, column=2, value=m).alignment = cls.ALIGN_CENTER
            c_tot = ws_dash.cell(row=r_i, column=3, value=f"=SUMIF('Budget Planning Detail'!$F$2:$F${max(2, total_row_idx-1)}, B{r_i}, 'Budget Planning Detail'!$J$2:$J${max(2, total_row_idx-1)})")
            c_real = ws_dash.cell(row=r_i, column=4, value=f"=SUMIFS('Budget Planning Detail'!$J$2:$J${max(2, total_row_idx-1)}, 'Budget Planning Detail'!$F$2:$F${max(2, total_row_idx-1)}, B{r_i}, 'Budget Planning Detail'!$K$2:$K${max(2, total_row_idx-1)}, \"Realized (PR Issued)\")")
            c_rem = ws_dash.cell(row=r_i, column=5, value=f"=C{r_i}-D{r_i}")
            for cell in [c_tot, c_real, c_rem]:
                cell.number_format = "#,##0"
                cell.alignment = cls.ALIGN_RIGHT
            for c_i in range(2, 6):
                c = ws_dash.cell(row=r_i, column=c_i)
                c.font = cls.REGULAR_FONT
                c.border = cls.THIN_BORDER_GRAY

        tot_m_row = 18 + len(month_names)
        ws_dash.cell(row=tot_m_row, column=2, value="TOTAL").alignment = cls.ALIGN_CENTER
        ws_dash.cell(row=tot_m_row, column=3, value=f"=SUM(C18:C{tot_m_row-1})").number_format = "#,##0"
        ws_dash.cell(row=tot_m_row, column=4, value=f"=SUM(D18:D{tot_m_row-1})").number_format = "#,##0"
        ws_dash.cell(row=tot_m_row, column=5, value=f"=SUM(E18:E{tot_m_row-1})").number_format = "#,##0"
        for c_i in range(2, 6):
            c = ws_dash.cell(row=tot_m_row, column=c_i)
            c.font = cls.BOLD_FONT
            c.fill = cls.SUBTOTAL_FILL
            c.border = cls.TOTAL_ROW_BORDER

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def generate_pr_excel(cls, periode: str) -> io.BytesIO:
        wb = openpyxl.Workbook()

        # Sheet 1: Procurement KPI Summary
        ws_dash = wb.active
        ws_dash.title = "Procurement KPI Summary"
        ws_dash.views.sheetView[0].showGridLines = True

        # Sheet 2: PR to Invoice Tracking
        ws_data = wb.create_sheet(title="PR to Invoice Tracking")
        ws_data.views.sheetView[0].showGridLines = True

        # Headers for PR Tracking Sheet
        proc_headers = [
            "No", "Requisition ID", "PR DocNum", "Request Date", "Req Category", "Line", "Part Number", 
            "Item Description", "PR Qty", "U/M", "Est. Unit Price (IDR)", "Total Est. PR Amount (IDR)",
            "PO ID", "PO Line", "PO Release", "PO DocNum", "PO Order Date", "Supplier Name", 
            "PO Qty", "PO Unit Price (IDR)", "Total PO Amount (IDR)", "GR Legal Number", "Packing Slip", 
            "Receipt Date", "Supplier Qty", "Invoice Number", "Invoice Date", "Comment / Remarks", 
            "PR Approval Status", "PO Approval Status", "Pipeline Stage"
        ]

        ws_data.row_dimensions[1].height = 28
        for col_idx, h in enumerate(proc_headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.ALIGN_HEADER
            cell.border = cls.TABLE_HEADER_BORDER

        # Fetch PRs & Kategoris
        kategoris = {k.id: k for k in Kategori.query.all()}
        prs = PrPoData.query.filter(PrPoData.status_ai != "CANCELLED").order_by(PrPoData.id.asc()).all()

        for i, pr in enumerate(prs, 1):
            row_idx = i + 1
            ws_data.row_dimensions[row_idx].height = 20

            kat = kategoris.get(pr.kategori_id)
            kat_kode = kat.kode if kat else ""

            if pr.invoice:
                stage = "4. Invoiced"
            elif pr.gr_legal_number:
                stage = "3. Goods Received"
            elif pr.po_doc_num:
                stage = "2. PO Issued"
            else:
                stage = "1. PR Only (Pending PO)"

            req_date_str = pr.request_date.strftime("%Y-%m-%d") if isinstance(pr.request_date, (datetime.datetime, datetime.date)) else (str(pr.request_date) if pr.request_date else "")
            order_date_str = pr.order_date.strftime("%Y-%m-%d") if isinstance(pr.order_date, (datetime.datetime, datetime.date)) else (str(pr.order_date) if pr.order_date else "")
            rcpt_date_str = pr.receipt_date.strftime("%Y-%m-%d") if isinstance(pr.receipt_date, (datetime.datetime, datetime.date)) else (str(pr.receipt_date) if pr.receipt_date else "")
            inv_date_str = pr.invoice_date.strftime("%Y-%m-%d") if isinstance(pr.invoice_date, (datetime.datetime, datetime.date)) else (str(pr.invoice_date) if pr.invoice_date else "")

            qty = float(pr.qty or 0.0)
            u_price = float(pr.unit_price or 0.0)
            t_price = float(pr.total_price or 0.0)

            c1 = ws_data.cell(row=row_idx, column=1, value=i)
            c2 = ws_data.cell(row=row_idx, column=2, value=pr.requisition_id or pr.id)
            c3 = ws_data.cell(row=row_idx, column=3, value=pr.pr_doc_num)
            c4 = ws_data.cell(row=row_idx, column=4, value=req_date_str)
            c5 = ws_data.cell(row=row_idx, column=5, value=kat_kode)
            c6 = ws_data.cell(row=row_idx, column=6, value=1)
            c7 = ws_data.cell(row=row_idx, column=7, value="")
            c8 = ws_data.cell(row=row_idx, column=8, value=pr.description)
            c9 = ws_data.cell(row=row_idx, column=9, value=qty)
            c10 = ws_data.cell(row=row_idx, column=10, value=pr.uom or "PCS")
            c11 = ws_data.cell(row=row_idx, column=11, value=u_price)
            c12 = ws_data.cell(row=row_idx, column=12, value=t_price if t_price else f"=I{row_idx}*K{row_idx}")

            c13 = ws_data.cell(row=row_idx, column=13, value=pr.po_doc_num or "")
            c14 = ws_data.cell(row=row_idx, column=14, value=1 if pr.po_doc_num else "")
            c15 = ws_data.cell(row=row_idx, column=15, value=1 if pr.po_doc_num else "")
            c16 = ws_data.cell(row=row_idx, column=16, value=pr.po_doc_num or "")
            c17 = ws_data.cell(row=row_idx, column=17, value=order_date_str)
            c18 = ws_data.cell(row=row_idx, column=18, value=pr.supplier_name or "")
            c19 = ws_data.cell(row=row_idx, column=19, value=qty if pr.po_doc_num else None)
            c20 = ws_data.cell(row=row_idx, column=20, value=u_price if pr.po_doc_num else None)
            c21 = ws_data.cell(row=row_idx, column=21, value=t_price if pr.po_doc_num else None)

            c22 = ws_data.cell(row=row_idx, column=22, value=pr.gr_legal_number or "")
            c23 = ws_data.cell(row=row_idx, column=23, value=pr.packing_slip or "")
            c24 = ws_data.cell(row=row_idx, column=24, value=rcpt_date_str)
            c25 = ws_data.cell(row=row_idx, column=25, value=qty if pr.gr_legal_number else None)

            c26 = ws_data.cell(row=row_idx, column=26, value=pr.invoice or "")
            c27 = ws_data.cell(row=row_idx, column=27, value=inv_date_str)
            c28 = ws_data.cell(row=row_idx, column=28, value=pr.comment_text or "")
            c29 = ws_data.cell(row=row_idx, column=29, value=pr.pr_status or "Approved")
            c30 = ws_data.cell(row=row_idx, column=30, value=pr.po_status or ("Approved" if pr.po_doc_num else ""))
            c31 = ws_data.cell(row=row_idx, column=31, value=stage)

            for col_c, align in [
                (c1, cls.ALIGN_CENTER), (c2, cls.ALIGN_CENTER), (c3, cls.ALIGN_CENTER), (c4, cls.ALIGN_CENTER),
                (c5, cls.ALIGN_CENTER), (c6, cls.ALIGN_CENTER), (c7, cls.ALIGN_CENTER), (c8, cls.ALIGN_LEFT),
                (c9, cls.ALIGN_RIGHT), (c10, cls.ALIGN_CENTER), (c11, cls.ALIGN_RIGHT), (c12, cls.ALIGN_RIGHT),
                (c13, cls.ALIGN_CENTER), (c14, cls.ALIGN_CENTER), (c15, cls.ALIGN_CENTER), (c16, cls.ALIGN_CENTER),
                (c17, cls.ALIGN_CENTER), (c18, cls.ALIGN_LEFT), (c19, cls.ALIGN_RIGHT), (c20, cls.ALIGN_RIGHT),
                (c21, cls.ALIGN_RIGHT), (c22, cls.ALIGN_CENTER), (c23, cls.ALIGN_CENTER), (c24, cls.ALIGN_CENTER),
                (c25, cls.ALIGN_RIGHT), (c26, cls.ALIGN_CENTER), (c27, cls.ALIGN_CENTER), (c28, cls.ALIGN_LEFT),
                (c29, cls.ALIGN_CENTER), (c30, cls.ALIGN_CENTER), (c31, cls.ALIGN_CENTER)
            ]:
                col_c.alignment = align
                col_c.font = cls.REGULAR_FONT
                col_c.border = cls.THIN_BORDER_GRAY

            c9.number_format = "#,##0.00"
            c11.number_format = "#,##0"
            c12.number_format = "#,##0"
            c19.number_format = "#,##0.00" if pr.po_doc_num else "General"
            c20.number_format = "#,##0" if pr.po_doc_num else "General"
            c21.number_format = "#,##0" if pr.po_doc_num else "General"
            c25.number_format = "#,##0.00" if pr.gr_legal_number else "General"

            if i % 2 == 0:
                for c in [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16, c17, c18, c19, c20, c21, c22, c23, c24, c25, c26, c27, c28, c29, c30, c31]:
                    c.fill = cls.ZEBRA_FILL

        total_row_idx = len(prs) + 2
        ws_data.row_dimensions[total_row_idx].height = 24
        c_lbl = ws_data.cell(row=total_row_idx, column=8, value="TOTAL PROCUREMENT AMOUNT")
        c_lbl.font = cls.BOLD_FONT
        c_lbl.alignment = cls.ALIGN_RIGHT

        c_tot_pr = ws_data.cell(row=total_row_idx, column=12, value=f"=SUM(L2:L{max(2, total_row_idx-1)})")
        c_tot_pr.font = cls.BOLD_FONT
        c_tot_pr.alignment = cls.ALIGN_RIGHT
        c_tot_pr.number_format = "#,##0"

        c_tot_po = ws_data.cell(row=total_row_idx, column=21, value=f"=SUM(U2:U{max(2, total_row_idx-1)})")
        c_tot_po.font = cls.BOLD_FONT
        c_tot_po.alignment = cls.ALIGN_RIGHT
        c_tot_po.number_format = "#,##0"

        for col in range(1, 32):
            cell = ws_data.cell(row=total_row_idx, column=col)
            cell.fill = cls.SUBTOTAL_FILL
            cell.border = cls.TOTAL_ROW_BORDER

        ws_data.freeze_panes = "A2"
        ws_data.auto_filter.ref = f"A1:AE{max(2, total_row_idx-1)}"

        # Column widths
        for col in ws_data.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                v = str(cell.value or "")
                if cell.number_format and "#,##0" in cell.number_format and isinstance(cell.value, (int, float)):
                    v = f"{cell.value:,.0f}"
                max_len = max(max_len, len(v))
            ws_data.column_dimensions[col_letter].width = max(max_len + 3, 10)
        ws_data.column_dimensions["H"].width = 42
        ws_data.column_dimensions["R"].width = 30
        ws_data.column_dimensions["AB"].width = 30

        # Build Procurement KPI Summary Sheet
        ws_dash.column_dimensions["A"].width = 3
        ws_dash.column_dimensions["B"].width = 24
        ws_dash.column_dimensions["C"].width = 18
        ws_dash.column_dimensions["D"].width = 18
        ws_dash.column_dimensions["E"].width = 18
        ws_dash.column_dimensions["F"].width = 4
        ws_dash.column_dimensions["G"].width = 24
        ws_dash.column_dimensions["H"].width = 18
        ws_dash.column_dimensions["I"].width = 18

        ws_dash.merge_cells("B2:I2")
        ws_dash["B2"] = f"PROCUREMENT & PR TRACKING SUMMARY - {periode}"
        ws_dash["B2"].font = cls.TITLE_FONT
        ws_dash.row_dimensions[2].height = 26

        ws_dash.merge_cells("B3:I3")
        ws_dash["B3"] = f"PT Summit Adyawinsa Indonesia | End-to-End PR to Invoice Pipeline Report"
        ws_dash["B3"].font = cls.SUBTITLE_FONT
        ws_dash.row_dimensions[3].height = 18

        kpis = [
            ("B5:C5", "B6:C6", "B7:C7", "TOTAL PR REQUISITIONS", f"=COUNTA('PR to Invoice Tracking'!C2:C{max(2, total_row_idx-1)})", f"{len(prs)} Item Line Items"),
            ("D5:E5", "D6:E6", "D7:E7", "TOTAL PR AMOUNT", f"=SUM('PR to Invoice Tracking'!L2:L{max(2, total_row_idx-1)})", "Est. Total PR Value"),
            ("G5:H5", "G6:H6", "G7:H7", "TOTAL PO ISSUED", f"=SUM('PR to Invoice Tracking'!U2:U{max(2, total_row_idx-1)})", "Realized Committed Spend"),
            ("I5:I5", "I6:I6", "I7:I7", "INVOICE FULFILLMENT", f"=COUNTIF('PR to Invoice Tracking'!AE2:AE{max(2, total_row_idx-1)}, \"4. Invoiced\")", "Fully Invoiced Items")
        ]

        for title_range, val_range, sub_range, title, formula, sub in kpis:
            t_start = title_range.split(":")[0]
            v_start = val_range.split(":")[0]
            s_start = sub_range.split(":")[0]

            if ":" in title_range and title_range.split(":")[0] != title_range.split(":")[1]:
                ws_dash.merge_cells(title_range)
                ws_dash.merge_cells(val_range)
                ws_dash.merge_cells(sub_range)

            ws_dash[t_start] = title
            ws_dash[t_start].font = cls.KPI_FONT_TITLE
            ws_dash[t_start].alignment = cls.ALIGN_CENTER
            ws_dash[t_start].fill = cls.KPI_BG_FILL

            ws_dash[v_start] = formula
            ws_dash[v_start].font = cls.KPI_FONT_VAL
            ws_dash[v_start].alignment = cls.ALIGN_CENTER
            ws_dash[v_start].fill = cls.KPI_BG_FILL
            if "PR" in title and "AMOUNT" in title or "PO" in title:
                ws_dash[v_start].number_format = "Rp #,##0"
            else:
                ws_dash[v_start].number_format = "#,##0"

            ws_dash[s_start] = sub
            ws_dash[s_start].font = cls.KPI_FONT_SUB
            ws_dash[s_start].alignment = cls.ALIGN_CENTER
            ws_dash[s_start].fill = cls.KPI_BG_FILL

        # Table 1: Pipeline Stage Breakdown
        ws_dash["B9"] = "1. REKAPITULASI TAHAPAN PIPELINE"
        ws_dash["B9"].font = cls.SECTION_FONT

        stage_headers = ["Tahapan Pipeline", "Jumlah PR", "Est. Total PR (IDR)"]
        for c_i, h in enumerate(stage_headers, 2):
            cell = ws_dash.cell(row=10, column=c_i, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.ALIGN_HEADER
            cell.border = cls.TABLE_HEADER_BORDER

        stages = [
            "1. PR Only (Pending PO)",
            "2. PO Issued",
            "3. Goods Received",
            "4. Invoiced"
        ]
        for r_i, st in enumerate(stages, 11):
            ws_dash.cell(row=r_i, column=2, value=st).alignment = cls.ALIGN_LEFT
            c_cnt = ws_dash.cell(row=r_i, column=3, value=f"=COUNTIF('PR to Invoice Tracking'!$AE$2:$AE${max(2, total_row_idx-1)}, B{r_i})")
            c_amt = ws_dash.cell(row=r_i, column=4, value=f"=SUMIF('PR to Invoice Tracking'!$AE$2:$AE${max(2, total_row_idx-1)}, B{r_i}, 'PR to Invoice Tracking'!$L$2:$L${max(2, total_row_idx-1)})")
            c_cnt.number_format = "#,##0"
            c_cnt.alignment = cls.ALIGN_RIGHT
            c_amt.number_format = "#,##0"
            c_amt.alignment = cls.ALIGN_RIGHT
            for c_i in range(2, 5):
                c = ws_dash.cell(row=r_i, column=c_i)
                c.font = cls.REGULAR_FONT
                c.border = cls.THIN_BORDER_GRAY

        ws_dash.cell(row=15, column=2, value="TOTAL").alignment = cls.ALIGN_LEFT
        ws_dash.cell(row=15, column=3, value="=SUM(C11:C14)").number_format = "#,##0"
        ws_dash.cell(row=15, column=4, value="=SUM(D11:D14)").number_format = "#,##0"
        for c_i in range(2, 5):
            c = ws_dash.cell(row=15, column=c_i)
            c.font = cls.BOLD_FONT
            c.fill = cls.SUBTOTAL_FILL
            c.border = cls.TOTAL_ROW_BORDER

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
