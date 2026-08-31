import os
import sys
import unittest
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import app
from services.excel.excel_export_service import ExcelExportService


class TestExcelExportService(unittest.TestCase):

    def setUp(self):
        self.app_context = app.app_context()
        self.app_context.push()

    def tearDown(self):
        self.app_context.pop()

    def test_generate_planning_excel(self):
        excel_stream = ExcelExportService.generate_planning_excel("2026")
        self.assertIsNotNone(excel_stream)
        self.assertGreater(excel_stream.getbuffer().nbytes, 0)

        # Parse with openpyxl
        wb = openpyxl.load_workbook(excel_stream, data_only=False)
        self.assertIn("Executive Summary", wb.sheetnames)
        self.assertIn("Budget Planning Detail", wb.sheetnames)

        ws_dash = wb["Executive Summary"]
        self.assertEqual(ws_dash["B2"].value, "EXECUTIVE SUMMARY - PLANNING BUDGET 2026")

        ws_data = wb["Budget Planning Detail"]
        self.assertEqual(ws_data["E1"].value, "Item Description")
        self.assertEqual(ws_data["J1"].value, "Planning Amount (IDR)")

    def test_generate_pr_excel(self):
        excel_stream = ExcelExportService.generate_pr_excel("2026")
        self.assertIsNotNone(excel_stream)
        self.assertGreater(excel_stream.getbuffer().nbytes, 0)

        # Parse with openpyxl
        wb = openpyxl.load_workbook(excel_stream, data_only=False)
        self.assertIn("Procurement KPI Summary", wb.sheetnames)
        self.assertIn("PR to Invoice Tracking", wb.sheetnames)

        ws_dash = wb["Procurement KPI Summary"]
        self.assertIn("PROCUREMENT & PR TRACKING SUMMARY", ws_dash["B2"].value)

        ws_data = wb["PR to Invoice Tracking"]
        self.assertEqual(ws_data["C1"].value, "PR DocNum")
        self.assertEqual(ws_data["H1"].value, "Item Description")


if __name__ == "__main__":
    unittest.main()
